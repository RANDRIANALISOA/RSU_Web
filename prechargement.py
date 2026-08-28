# -*- coding: utf-8 -*-
"""
prechargement.py — Génération web de la BASE DE PRÉCHARGEMENT (VAD).

Portage WEB de `..\\RSU_Rapport\\base_prechargement.py` : produit le même classeur
Excel **base_prechargement_<version>.xlsx** (3 feuilles : Ensemble / nouveau /
e_fokontany) ET le fichier **charge_agents_<version>.xlsx** (récapitulatif de la
charge par agent), avec le MÊME équilibrage de charge (`affectation_agents.py`,
copié depuis le projet exe). Les deux fichiers sont renvoyés dans une archive ZIP.

DIFFÉRENCE ESSENTIELLE avec la version exe : les données ne sont pas lues dans des
`.dta` mais dans la **base SQL** (via `db_source.DbDataset`), filtrées sur le
district d'affectation de l'Expert « Traitement ». Comme la base web ne contient pas
`base_login_ce.dta` / `base_login_enq.dta`, les colonnes CE / ENQ sont DÉRIVÉES :

  • ENQ (agent enquêteur) = `interview__diagnostics.responsible` (CODE agent),
    joint aux ménages par `interview__key` ;
  • CE (chef d'équipe) = le CODE du chef de cet agent via les tables `equipes`
    (`agent.login_ce`). CE et ENQ portent TOUJOURS des CODES (le nom reste le
    libellé côté base / rapport, jamais dans ces colonnes).

L'unité indéplaçable de l'équilibrage reste l'`interview__key` (une visite de
dénombrement = un seul agent, ~quelques dizaines de ménages), exactement comme dans
le programme exe. Après équilibrage, l'ENQ change mais le CE d'origine est conservé.

Bibliothèque standard + openpyxl (déjà utilisé partout dans le web).
"""
from __future__ import annotations

import io
import statistics
import zipfile
from collections import Counter
from datetime import datetime

import db_source
import equipes
from affectation_agents import equilibrer_charge
from rapport_core import ErreurDonnees, _log_noop


# ---------------------------------------------------------------------------
# Colonnes exportées par feuille (ordre = fichier de référence, cf. exe)
# ---------------------------------------------------------------------------
COLS_ENSEMBLE = [
    "interview__key", "region", "district", "commune", "fokontany",
    "fkt_recherche", "nom_cm", "IdeFKT", "taille_men", "adresse", "description",
    "gps_coord__Latitude", "gps_coord__Longitude", "code_den",
    "CE", "ENQ", "interview_keyden",
]
COLS_EFOKONTANY = COLS_ENSEMBLE
COLS_NOUVEAU = [
    "interview__key", "region", "district", "commune", "fokontany",
    "fkt_recherche", "nom_cm", "taille_men", "CQ17_preload", "description",
    "GPS_Lat_ZD", "GPS_Long_ZD", "code_den", "_responsible", "_quantity",
    "IdeFKT_recherche", "typemen", "mode_enreg", "nom_projet", "interview_keyden",
]

# Modes d'affectation acceptés (mêmes clés que le programme exe).
MODES = ("denombrement", "equilibre", "equilibre_fort")
LIBELLE_MODE = {
    "denombrement": "Dénombrement (aucune redistribution)",
    "equilibre": "Équilibré (±10 %, agents groupés)",
    "equilibre_fort": "Équilibré fort (échanges ; un agent peut changer de fokontany)",
}


# ---------------------------------------------------------------------------
# Helpers d'apurement (identiques au programme exe base_prechargement.py)
# ---------------------------------------------------------------------------
def _est_na(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() in ("", "##N/A##"))


def _sans_na(v) -> str:
    return "" if _est_na(v) else str(v)


def _nettoyer_idefkt(raw) -> str:
    """ID_efkt -> code du carnet e-fokontany, ou '' (reproduit le dofile)."""
    s = "" if raw is None else str(raw)
    if len(s) > 80:
        i = s.find("|")
        if i >= 0:
            s = s[i + 1:].strip()
    if len(s) < 15:
        s = ""
    return s


def _gps_num(v):
    """Coordonnée GPS -> float exploitable, ou None (manquante/nulle)."""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if f == 0 else f


def _gps_txt(v) -> str:
    """Coordonnée GPS -> texte ('' si manquante ou nulle)."""
    if v is None:
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return _sans_na(v)
    if f == 0:
        return ""
    return repr(round(f, 8))


def _codes_hierarchie(code8, reg_brut, dis_brut, com_brut):
    """Code fokontany (8 chiffres) -> (region, district, commune, fokontany)."""
    try:
        c = int(code8)
    except (TypeError, ValueError):
        c = None
    if c and c >= 1_000_000:
        return (c // 1_000_000, c // 10_000, c // 100, c)
    return (reg_brut, dis_brut, com_brut, code8)


# ---------------------------------------------------------------------------
# Exclusion des ménages déjà envoyés (préchargements précédents téléversés)
# ---------------------------------------------------------------------------
def _keyden_precedents(fichiers, log) -> set:
    """`fichiers` = liste de (nom, octets) de classeurs .xlsx déjà générés.
    Renvoie l'ensemble des `interview_keyden` déjà envoyés (à exclure). Un fichier
    illisible est ignoré sans bloquer."""
    import openpyxl

    deja: set = set()
    for nom, octets in fichiers:
        if not octets:
            continue
        try:
            wb = openpyxl.load_workbook(io.BytesIO(octets), read_only=True,
                                        data_only=True)
        except Exception:
            log(f"   Fichier ignore (illisible) : {nom}")
            continue
        n0 = len(deja)
        for ws in wb.worksheets:
            lignes = ws.iter_rows(values_only=True)
            try:
                entete = next(lignes)
            except StopIteration:
                continue
            try:
                j = list(entete).index("interview_keyden")
            except ValueError:
                continue
            for row in lignes:
                if j < len(row) and row[j] not in (None, ""):
                    deja.add(str(row[j]))
        wb.close()
        log(f"   {nom} : {len(deja) - n0} nouveaux menages deja envoyes")
    log(f"   Total menages deja envoyes (a exclure) : {len(deja)}")
    return deja


# ---------------------------------------------------------------------------
# Lecture des ménages depuis la base (district) + apurement
# ---------------------------------------------------------------------------
def _charger_menages(conn, code_district: int, log):
    """Reconstruit (ensemble, nouveau, efokontany) pour un district, depuis la base.

    ENQ = code agent (interview__diagnostics.responsible) ; CE = chef d'équipe via
    les tables `equipes`. Reproduit fidèlement l'apurement du programme exe."""
    src = db_source.source_db(conn, district=code_district)
    den = src("den")
    ros = src("roster")
    diag = src("diagnostics")

    log("[1/4] Lecture de DEN_MENAGE (zones geographiques)...")
    den_keys = den.col("interview__key")
    den_reg = den.col("region")
    den_dis = den.col("district")
    den_com = den.col("commune")
    den_fkt = den.col("fokontany")
    den_fkt_dec = den.col_decoded("fokontany")
    den_num = den.col("num_fkt")
    den_seg = den.col("segment")
    den_index = {}
    for i, k in enumerate(den_keys):
        if k not in den_index:
            code8 = den_fkt[i] if den_fkt[i] is not None else den_num[i]
            den_index[k] = {
                "region": den_reg[i], "district": den_dis[i],
                "commune": den_com[i], "code8": code8,
                "fkt_recherche": den_fkt_dec[i],
                "segment": "" if den_seg[i] is None else str(den_seg[i]),
            }

    log("[2/4] Derivation CE / ENQ (diagnostics + equipes)...")
    # ENQ (agent enquêteur) par interview__key = interview__diagnostics.responsible.
    d_keys = diag.col("interview__key")
    d_resp = diag.col("responsible")
    enq_index = {}
    for i, k in enumerate(d_keys):
        if k not in enq_index:
            enq_index[k] = _sans_na(d_resp[i])
    # CE (chef d'équipe) : chef de l'agent, via les tables equipes.
    liens = equipes.agents_et_chefs(conn)   # {code_agent: {nom, chef_login, chef_nom}}
    log(f"   ENQ : {len(enq_index)} interviews ; agents connus : {len(liens)}")

    log("[3/4] Lecture de segment_roster (menages) + apurement...")
    r_key = ros.col("interview__key")
    r_srid = ros.col("segment_roster__id")
    r_nom = ros.col("nom_cm")
    r_nomD = ros.col("nom_cmD")
    r_surnom = ros.col("surnom")
    r_taille = ros.col("taille_menD")
    r_adr = ros.col("adresse")
    r_desc = ros.col("description")
    r_lat = ros.col("gps_coord__Latitude")
    r_lon = ros.col("gps_coord__Longitude")
    r_code = ros.col("code_den")
    r_idefkt = ros.col("ID_efkt")

    ensemble, nouveau, efokontany = [], [], []
    vus = set()
    n_drop_vide = n_drop_code = n_drop_nomerge = 0
    for i in range(ros.nobs):
        k = r_key[i]
        info = den_index.get(k)
        if info is None:                    # pas dans DEN_MENAGE (_merge != 3)
            n_drop_nomerge += 1
            continue
        if (r_nomD[i] == "##N/A##" and r_surnom[i] == "##N/A##"
                and r_taille[i] is None and r_adr[i] == "##N/A##"):
            n_drop_vide += 1
            continue
        code_den = _sans_na(r_code[i])
        if code_den == "//()":
            n_drop_code += 1
            continue
        srid = "" if r_srid[i] is None else str(int(r_srid[i]))
        keyden = k + "-" + srid.zfill(3)
        if keyden in vus:
            continue
        vus.add(keyden)

        reg, dis, com, fkt = _codes_hierarchie(
            info["code8"], info["region"], info["district"], info["commune"])
        idefkt = _nettoyer_idefkt(r_idefkt[i])
        est_nouveau = (idefkt == "")
        adresse = "" if _est_na(r_adr[i]) else str(r_adr[i]).upper()
        lat, lon = _gps_txt(r_lat[i]), _gps_txt(r_lon[i])
        lat_n, lon_n = _gps_num(r_lat[i]), _gps_num(r_lon[i])
        enq = enq_index.get(k, "")
        ce = ""
        lien = liens.get(enq)
        if lien:
            ce = lien.get("chef_login") or ""    # CODE du chef d'équipe (login_ce)

        base = {
            "interview__key": k,
            "interview_keyden": keyden,
            "region": reg, "district": dis, "commune": com, "fokontany": fkt,
            "fkt_recherche": _sans_na(info["fkt_recherche"]),
            "nom_cm": _sans_na(r_nom[i]),
            "taille_men": "" if r_taille[i] is None else r_taille[i],
            "description": _sans_na(r_desc[i]),
            "code_den": code_den,
            "CE": ce, "ENQ": enq,
            # Méta internes (préfixe _) : équilibrage + charge par agent.
            "_sid": k, "_seg": info["segment"], "_com": com,
            "_fktnom": _sans_na(info["fkt_recherche"]) or str(fkt),
            "_lat": lat_n, "_lon": lon_n, "_agent": enq,
        }
        ligne_ens = {**base, "IdeFKT": idefkt, "adresse": adresse,
                     "gps_coord__Latitude": lat, "gps_coord__Longitude": lon}
        ensemble.append(ligne_ens)
        if est_nouveau:
            nouveau.append({**base, "CQ17_preload": adresse,
                            "GPS_Lat_ZD": lat, "GPS_Long_ZD": lon,
                            "_responsible": enq, "_quantity": 1,
                            "IdeFKT_recherche": "", "typemen": 2,
                            "mode_enreg": 1, "nom_projet": 1})
        else:
            efokontany.append(ligne_ens)

    log(f"   {len(ensemble)} menages : {len(efokontany)} e-fokontany, "
        f"{len(nouveau)} nouveaux (ignores : {n_drop_nomerge} sans segment, "
        f"{n_drop_vide} vides, {n_drop_code} code invalide)")
    return ensemble, nouveau, efokontany


# ---------------------------------------------------------------------------
# Équilibrage de la charge (délègue à affectation_agents, comme l'exe)
# ---------------------------------------------------------------------------
def _appliquer_affectation(ensemble, nouveau, efokontany, mode: str, log):
    if mode not in ("equilibre", "equilibre_fort"):
        log("Affectation : agent du denombrement (aucune redistribution).")
        return
    fort = (mode == "equilibre_fort")
    if fort:
        log("Equilibrage FORT de la charge (par commune, avec echanges)...")
    else:
        log("Equilibrage de la charge entre agents (par commune)...")
    agg: dict = {}
    for r in ensemble:
        u = agg.get(r["_sid"])
        if u is None:
            u = agg[r["_sid"]] = {
                "sid": r["_sid"], "commune": r["_com"], "n": 0,
                "agent": r["_agent"], "_slat": 0.0, "_slon": 0.0, "_w": 0}
        u["n"] += 1
        if r["_lat"] is not None and r["_lon"] is not None:
            u["_slat"] += r["_lat"]
            u["_slon"] += r["_lon"]
            u["_w"] += 1
    unites = []
    for u in agg.values():
        u["lat"] = u["_slat"] / u["_w"] if u["_w"] else None
        u["lon"] = u["_slon"] / u["_w"] if u["_w"] else None
        unites.append(u)
    affect = equilibrer_charge(unites, log,
                               tol_frac=(0.0 if fort else 0.10), echanges=fort)
    n_moves = 0
    for r in ensemble:
        na = affect.get(r["_sid"], r["_agent"])
        if na != r["ENQ"]:
            n_moves += 1
        r["ENQ"] = na
    for r in efokontany:
        r["ENQ"] = affect.get(r["_sid"], r["_agent"])
    for r in nouveau:
        r["_responsible"] = affect.get(r["_sid"], r["_agent"])
    log(f"   {n_moves} menage(s) reaffecte(s) a un autre agent "
        f"(le CE d'origine est conserve).")


# ---------------------------------------------------------------------------
# Écriture des deux classeurs -> octets
# ---------------------------------------------------------------------------
def _octets_prechargement(ensemble, nouveau, efokontany) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook(write_only=True)
    for titre, cols, lignes in (
            ("Ensemble", COLS_ENSEMBLE, ensemble),
            ("nouveau", COLS_NOUVEAU, nouveau),
            ("e_fokontany", COLS_EFOKONTANY, efokontany)):
        ws = wb.create_sheet(titre)
        ws.append(cols)
        for ligne in lignes:
            ws.append([ligne.get(c, "") for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stats(valeurs):
    """min, max, médiane, moyenne, écart-type (arrondis) d'une liste."""
    if not valeurs:
        return (0, 0, 0, 0, 0)
    mn, mx = min(valeurs), max(valeurs)
    med = round(statistics.median(valeurs), 1)
    moy = round(statistics.fmean(valeurs), 1)
    ect = round(statistics.pstdev(valeurs), 1) if len(valeurs) > 1 else 0
    return (mn, mx, med, moy, ect)


def _octets_charge(ensemble) -> bytes:
    """Fichier de charge par agent (reflète l'affectation finale, colonne ENQ)."""
    import openpyxl
    par_agent: Counter = Counter()
    par_segment: Counter = Counter()
    agent_commune: dict = {}
    for r in ensemble:
        agent_final = r["ENQ"]
        com = r["_com"]
        par_agent[agent_final] += 1
        par_segment[(com, r["_fktnom"], r["_seg"], agent_final)] += 1
        agent_commune.setdefault(agent_final, com)

    wbc = openpyxl.Workbook(write_only=True)
    ws1 = wbc.create_sheet("Par_agent")
    ws1.append(["code_agent", "commune", "nb_menages"])
    for agent_final in sorted(par_agent):
        ws1.append([agent_final, agent_commune.get(agent_final, ""),
                    par_agent[agent_final]])
    ws2 = wbc.create_sheet("Par_segment")
    ws2.append(["commune", "fokontany", "segment", "code_agent", "nb_menages"])
    for cle in sorted(par_segment):
        ws2.append([cle[0], cle[1], cle[2], cle[3], par_segment[cle]])

    ws3 = wbc.create_sheet("Statistiques")
    charges = list(par_agent.values())
    total_men = sum(charges)
    nb_agents = len(par_agent)
    nb_com = len(set(agent_commune.values()))
    g_mn, g_mx, g_med, g_moy, g_ect = _stats(charges)
    ws3.append(["Statistiques globales (charge = nb de menages par agent)"])
    ws3.append(["Indicateur", "Valeur"])
    ws3.append(["Nombre d'agents", nb_agents])
    ws3.append(["Nombre de communes", nb_com])
    ws3.append(["Nombre total de menages", total_men])
    ws3.append(["Minimum (menages/agent)", g_mn])
    ws3.append(["Maximum (menages/agent)", g_mx])
    ws3.append(["Mediane (menages/agent)", g_med])
    ws3.append(["Moyenne (menages/agent)", g_moy])
    ws3.append(["Ecart-type (menages/agent)", g_ect])
    ws3.append([])

    charges_par_com: dict = {}
    menages_par_com: Counter = Counter()
    for agent_final, com in agent_commune.items():
        charges_par_com.setdefault(com, []).append(par_agent[agent_final])
    for r in ensemble:
        menages_par_com[r["_com"]] += 1
    ws3.append(["Repartition des agents et de la charge par commune"])
    ws3.append(["Commune", "Nombre d'agents", "Nombre de menages",
                "Min", "Max", "Mediane", "Moyenne", "Ecart-type"])
    for com in sorted(charges_par_com):
        vals = charges_par_com[com]
        c_mn, c_mx, c_med, c_moy, c_ect = _stats(vals)
        ws3.append([com, len(vals), menages_par_com[com],
                    c_mn, c_mx, c_med, c_moy, c_ect])
    ws3.append([])

    tranches = [(0, 50), (51, 100), (101, 150), (151, 200),
                (201, 300), (301, 500), (501, 10**9)]
    libelles = ["1 - 50", "51 - 100", "101 - 150", "151 - 200",
                "201 - 300", "301 - 500", "501 et +"]
    hist = [0] * len(tranches)
    for ch in charges:
        for idx, (lo, hi) in enumerate(tranches):
            if lo <= ch <= hi:
                hist[idx] += 1
                break
    ws3.append(["Repartition des agents par tranche de charge"])
    ws3.append(["Tranche (menages)", "Nombre d'agents"])
    for lib, nb in zip(libelles, hist):
        ws3.append([lib, nb])

    buf = io.BytesIO()
    wbc.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Point d'entrée web : génère les deux fichiers dans une archive ZIP
# ---------------------------------------------------------------------------
def generer_zip(conn, code_district: int, mode: str = "denombrement",
                exclure_fichiers=None, log=None) -> tuple:
    """Génère la base de préchargement + la charge par agent pour un district.

    Renvoie (nom_zip, octets_zip). `mode` ∈ MODES. `exclure_fichiers` = liste de
    (nom, octets) de préchargements .xlsx précédents dont les ménages sont exclus.
    Lève ErreurDonnees s'il n'y a aucun ménage exploitable."""
    log = log or _log_noop
    mode = (mode or "denombrement").strip().lower()
    if mode not in MODES:
        mode = "denombrement"

    ensemble, nouveau, efokontany = _charger_menages(conn, int(code_district), log)
    if not ensemble:
        raise ErreurDonnees(
            "Aucun menage exploitable pour ce district : rien a exporter. "
            "Le denombrement a-t-il ete transcrit dans la base ?")

    if exclure_fichiers:
        log("Exclusion des menages deja envoyes (prechargements precedents)...")
        deja = _keyden_precedents(exclure_fichiers, log)
        if deja:
            avant = len(ensemble)
            ensemble = [r for r in ensemble if r["interview_keyden"] not in deja]
            nouveau = [r for r in nouveau if r["interview_keyden"] not in deja]
            efokontany = [r for r in efokontany
                          if r["interview_keyden"] not in deja]
            log(f"   {avant - len(ensemble)} menage(s) deja envoye(s) retire(s) "
                f"-> {len(ensemble)} restant(s).")
            if not ensemble:
                raise ErreurDonnees(
                    "Tous les menages ont deja ete envoyes dans les prechargements "
                    "precedents : rien de nouveau a exporter.")

    log("[4/4] Affectation des agents + ecriture des classeurs...")
    _appliquer_affectation(ensemble, nouveau, efokontany, mode, log)

    version = datetime.now().strftime("%Y%m%d_%HH%MMN")
    oct_prech = _octets_prechargement(ensemble, nouveau, efokontany)
    oct_charge = _octets_charge(ensemble)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"base_prechargement_{version}.xlsx", oct_prech)
        z.writestr(f"charge_agents_{version}.xlsx", oct_charge)
    return (f"prechargement_district{code_district}_{version}.zip", buf.getvalue())


# ---------------------------------------------------------------------------
# Pages (charte visuelle d'admin.py / equipes.py)
# ---------------------------------------------------------------------------
import admin  # noqa: E402  (après les fonctions : évite un cycle au chargement)

ESC = admin.ESC


def page_prechargement(district_txt, erreur=None) -> str:
    """Formulaire : choix du mode d'affectation + exclusion optionnelle + génération."""
    h = [equipes._entete(),
         '<p style="margin:0 0 6px"><a href="/traitement">'
         '&larr; Accueil Traitement</a></p>',
         '<h1>Base de pr&eacute;chargement (VAD)</h1>',
         f'<div class="note">District d&rsquo;affectation : <b>{ESC(district_txt)}</b>. '
         'La base est construite &agrave; partir du <b>d&eacute;nombrement transcrit</b> '
         'de ce district. Deux fichiers Excel sont produits (dans une archive ZIP) : '
         'la <b>base de pr&eacute;chargement</b> (3 feuilles : Ensemble / nouveau / '
         'e_fokontany) et la <b>charge par agent</b> (r&eacute;capitulatif de '
         'l&rsquo;affectation).</div>']
    if erreur:
        h.append(f'<div class="err">{erreur}</div>')

    opts = "".join(
        f'<label style="display:block;margin:6px 0">'
        f'<input type="radio" name="mode" value="{m}"'
        f'{" checked" if m == "denombrement" else ""}> {ESC(LIBELLE_MODE[m])}</label>'
        for m in MODES)

    h.append(
        '<h2>G&eacute;n&eacute;rer</h2>'
        '<form method="post" action="/traitement/prechargement" '
        'enctype="multipart/form-data" class="grid-form">'
        '<div><label><b>Mode d&rsquo;affectation des m&eacute;nages aux agents</b></label>'
        f'{opts}</div>'
        '<div><label>Exclure les m&eacute;nages d&eacute;j&agrave; envoy&eacute;s '
        '(facultatif) &mdash; t&eacute;l&eacute;versez les fichiers '
        '<b>base_prechargement_*.xlsx</b> pr&eacute;c&eacute;demment g&eacute;n&eacute;r&eacute;s</label>'
        '<input type="file" name="exclure" accept=".xlsx" multiple></div>'
        '<div style="align-self:end"><button>G&eacute;n&eacute;rer et '
        't&eacute;l&eacute;charger (ZIP)</button></div>'
        '</form>')
    h.append('<div class="note">Modes : <b>D&eacute;nombrement</b> = chaque '
             'm&eacute;nage reste &agrave; l&rsquo;agent qui l&rsquo;a d&eacute;nombr&eacute; '
             '(aucune redistribution). <b>&Eacute;quilibr&eacute;</b> = r&eacute;partit la '
             'charge (&plusmn;10 %) en gardant les agents group&eacute;s '
             'g&eacute;ographiquement, sans changer de commune. <b>&Eacute;quilibr&eacute; '
             'fort</b> = pousse l&rsquo;&eacute;quilibre par des &eacute;changes de '
             'segments (un agent peut alors changer de fokontany).</div>')
    h.append('</div></body></html>')
    return "".join(h)
