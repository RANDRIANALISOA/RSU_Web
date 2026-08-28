"""export_rapport.py — Export Excel du rapport de dénombrement (par district).

Produit un classeur .xlsx à partir des données en base (via le moteur `rapport_core`
pour construire la liste des ménages, exactement comme le rapport HTML) :

  Feuille 1 « Rapport global » :
    - Couverture par commune vs projection RGPH-3 2025 (+ total district) ;
    - Qualité par commune : dénombrés, avec/sans carnet, scanné, % GPS (+ total) ;
    - Un tableau par commune : détail par fokontany (mêmes colonnes).
  Feuille 2 « Dénombrement par agent-jour » :
    - Ménages dénombrés par agent et par jour, regroupés par commune puis par chef
      d'équipe ; une colonne par date.

Codage carnet (cf. gabarit) : 1 = avec carnet (scanné), 2 = avec carnet (non
scanné), 3 = sans carnet, 4 = ne sait pas. Donc avec carnet = {1,2}, sans = {3},
scanné = {1}. GPS capturé = latitude ET longitude numériques.
"""

import datetime
import io

import db_source
import rapport_core
import zones
import equipes

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Styles ---------------------------------------------------------------
# NB : couleurs en ARGB 8 chiffres avec alpha OPAQUE « FF ». Sans le préfixe,
# openpyxl stocke un alpha « 00 » (transparent) que certains lecteurs (LibreOffice…)
# respectent -> fond « transparent » (blanc) + police blanche = texte invisible.
_TITRE = Font(bold=True, size=14, color="FF1F2937")
_SOUS = Font(bold=True, size=11, color="FF374151")
# En-tête : texte FONCÉ sur fond CLAIR (et non blanc sur bleu). Ainsi, même si un
# lecteur n'applique pas le remplissage, le texte reste lisible (jamais blanc/blanc).
_ENTETE_F = Font(bold=True, color="FF1F2937")
_ENTETE_FILL = PatternFill("solid", fgColor="FFDBEAFE")
_TOTAL_F = Font(bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="FFE5E7EB")
_COMMUNE_FILL = PatternFill("solid", fgColor="FFEFF6FF")
_bord = Side(style="thin", color="FFD1D5DB")
_BORDURE = Border(left=_bord, right=_bord, top=_bord, bottom=_bord)
_CENTRE = Alignment(horizontal="center")
_DROITE = Alignment(horizontal="right")


def _has_gps(m):
    la, lo = m.get("lat"), m.get("lon")
    return isinstance(la, (int, float)) and isinstance(lo, (int, float))


def _stats(ms):
    """(dénombrés, avec_carnet, sans_carnet, scanné, %GPS) pour une liste de ménages."""
    total = len(ms)
    avec = sum(1 for m in ms if m.get("carnet") in (1, 2))
    sans = sum(1 for m in ms if m.get("carnet") == 3)
    scan = sum(1 for m in ms if m.get("carnet") == 1)
    gps = sum(1 for m in ms if _has_gps(m))
    pct = round(100.0 * gps / total, 1) if total else 0.0
    return total, avec, sans, scan, pct


def _menages_scope(conn, code_district, communes_autorisees):
    """Construit la liste des ménages du périmètre via le moteur (codes agent bruts)."""
    if communes_autorisees:
        source = db_source.source_db(
            conn, communes=[int(c) for c in communes_autorisees])
    else:
        source = db_source.source_db(conn, district=code_district)
    diag = rapport_core._charger_diagnostics(source("diagnostics"), agents_noms=None)
    seg_by_key, _ = rapport_core._charger_segments(source("den"), diag)
    return rapport_core._construire_menages(source("roster"), seg_by_key)


def _ecrire_entete(ws, ligne, colonnes, largeurs=None):
    for j, titre in enumerate(colonnes, start=1):
        c = ws.cell(row=ligne, column=j, value=titre)
        c.font = _ENTETE_F
        c.fill = _ENTETE_FILL
        c.border = _BORDURE
        # Retour à la ligne : les intitulés longs s'affichent en entier.
        c.alignment = Alignment(
            horizontal=("center" if j > 1 else "left"),
            vertical="center", wrap_text=True)
    ws.row_dimensions[ligne].height = 42
    if largeurs:
        for j, w in enumerate(largeurs, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return ligne + 1


def _ligne(ws, ligne, valeurs, total=False, fill=None):
    for j, v in enumerate(valeurs, start=1):
        c = ws.cell(row=ligne, column=j, value=v)
        c.border = _BORDURE
        if j > 1 and isinstance(v, (int, float)):
            c.alignment = _DROITE
        if total:
            c.font = _TOTAL_F
        if fill:
            c.fill = fill
        elif total:
            c.fill = _TOTAL_FILL
    return ligne + 1


_SOURCE_F = Font(italic=True, size=9, color="FF6B7280")


def _titre(ws, ligne, texte):
    """Titre de tableau (gras) suivi d'une LIGNE VIDE. Renvoie la ligne de
    l'en-tête à écrire ensuite (titre en `ligne`, vide en `ligne+1`)."""
    ws.cell(row=ligne, column=1, value=texte).font = _SOUS
    return ligne + 2


def _sous_titre(ws, ligne, texte, ncols):
    """Titre de sous-section (habillé d'un fond clair) + ligne vide."""
    c = ws.cell(row=ligne, column=1, value=texte)
    c.font = _SOUS
    for j in range(1, ncols + 1):
        ws.cell(row=ligne, column=j).fill = _COMMUNE_FILL
    return ligne + 2


def _source(ws, ligne):
    """« Source » en ITALIQUE sous un tableau + une ligne vide. Renvoie la ligne
    suivante libre."""
    ws.cell(row=ligne, column=1,
            value="Source : Dénombrement RSU 2026").font = _SOURCE_F
    return ligne + 2


def _entete_document(ws, nom_district, note):
    """Bloc d'en-tête IDENTIQUE sur toutes les feuilles (parité de style avec la
    feuille « Rapport global ») : titre, district, date de génération, puis une
    note de lecture propre à la feuille (en italique). Renvoie la 1re ligne de
    contenu (6, avec une ligne vide en 5)."""
    ws.cell(row=1, column=1, value="Dénombrement RSU 2026").font = _TITRE
    ws.cell(row=2, column=1, value=f"District : {nom_district}").font = _SOUS
    ws.cell(row=3, column=1,
            value="Généré le : "
            + datetime.datetime.now().strftime("%d/%m/%Y %H:%M"))
    ws.cell(row=4, column=1, value=note).font = _SOURCE_F
    return 6


# ---------------------------------------------------------------------------
# Feuille 1 : Rapport global
# ---------------------------------------------------------------------------
def _feuille_global(wb, conn, code_district, nom_district, menages,
                    communes_autorisees):
    ws = wb.active
    ws.title = "Rapport global"

    # Référentiel : communes et fokontany du périmètre (inclut les vides).
    ref = zones.reference_district(conn, code_district)
    if communes_autorisees is not None:
        ref = [z for z in ref if z["ccode"] in communes_autorisees]
    communes = {}       # ccode -> nom
    fkt_par_commune = {}  # ccode -> [(fkt, label)]
    for z in ref:
        communes.setdefault(z["ccode"], z["commune"])
        fkt_par_commune.setdefault(z["ccode"], []).append((z["fkt"], z["label"]))
    ordre_com = sorted(communes, key=lambda c: (communes[c] or "", c))

    # Ménages regroupés par commune (fktcode[:6]) et par fokontany (fktcode).
    par_com, par_fkt = {}, {}
    for m in menages:
        fc = m.get("fktcode") or ""
        cc = fc[:6] if len(fc) >= 6 else ""
        par_com.setdefault(cc, []).append(m)
        par_fkt.setdefault(fc, []).append(m)

    # -- En-tête du document (bloc partagé) --
    r = _entete_document(
        ws, nom_district,
        "Lecture : sauf indication « (%) », toutes les valeurs sont des "
        "NOMBRES DE MÉNAGES. « Ménages attendus » = projection RGPH-3 2025 ; "
        "« Ménages dénombrés » = ménages recensés (RSU 2026).")

    # -- Tableau 1 : couverture par commune (projection RGPH-3 2025) --
    r = _titre(ws, r, "Tableau 1 — Couverture du dénombrement "
               "(ménages dénombrés vs projection RGPH-3 2025)")
    r = _ecrire_entete(
        ws, r,
        ["Commune", "Ménages attendus (projection RGPH-3 2025)",
         "Ménages dénombrés (RSU 2026)", "Taux de couverture (%)"],
        largeurs=[34, 30, 26, 20])
    attendus = zones.attendus_communes(conn, ordre_com)
    tot_att = tot_real = 0
    for cc in ordre_com:
        att = (attendus.get(cc, {}).get("attendu") or 0)
        real = len(par_com.get(cc, []))
        tot_att += att
        tot_real += real
        pct = round(100.0 * real / att, 1) if att else None
        r = _ligne(ws, r, [communes[cc], att, real,
                           (pct if pct is not None else "—")])
    pct_tot = round(100.0 * tot_real / tot_att, 1) if tot_att else None
    r = _ligne(ws, r, ["TOTAL DISTRICT", tot_att, tot_real,
                       (pct_tot if pct_tot is not None else "—")], total=True)
    r = _source(ws, r)

    # -- Tableau 2 : qualité par commune --
    r = _titre(ws, r, "Tableau 2 — Qualité du dénombrement par commune")
    r = _ecrire_entete(
        ws, r,
        ["Commune", "Ménages dénombrés", "dont avec carnet",
         "dont sans carnet", "dont carnet scanné", "GPS capturé (%)"],
        largeurs=[34, 18, 16, 16, 18, 16])
    T = [0, 0, 0, 0]
    tot_gps = 0
    for cc in ordre_com:
        ms = par_com.get(cc, [])
        total, avec, sans, scan, pctg = _stats(ms)
        T[0] += total; T[1] += avec; T[2] += sans; T[3] += scan
        tot_gps += sum(1 for m in ms if _has_gps(m))
        r = _ligne(ws, r, [communes[cc], total, avec, sans, scan, pctg])
    pg = round(100.0 * tot_gps / T[0], 1) if T[0] else 0.0
    r = _ligne(ws, r, ["TOTAL DISTRICT", T[0], T[1], T[2], T[3], pg], total=True)
    r = _source(ws, r)

    # -- Tableau 3 : détail par commune (par fokontany) --
    r = _titre(ws, r, "Tableau 3 — Détail par commune (par fokontany)")
    for cc in ordre_com:
        r = _sous_titre(ws, r, f"Commune : {communes[cc]}", 6)
        r = _ecrire_entete(
            ws, r,
            ["Fokontany", "Ménages dénombrés", "dont avec carnet",
             "dont sans carnet", "dont carnet scanné", "GPS capturé (%)"],
            largeurs=[34, 18, 16, 16, 18, 16])
        Tc = [0, 0, 0, 0]
        gps_c = 0
        for fkt, label in sorted(fkt_par_commune.get(cc, []),
                                 key=lambda x: (x[1] or "", x[0])):
            ms = par_fkt.get(fkt, [])
            total, avec, sans, scan, pctg = _stats(ms)
            Tc[0] += total; Tc[1] += avec; Tc[2] += sans; Tc[3] += scan
            gps_c += sum(1 for m in ms if _has_gps(m))
            r = _ligne(ws, r, [label, total, avec, sans, scan, pctg])
        pgc = round(100.0 * gps_c / Tc[0], 1) if Tc[0] else 0.0
        r = _ligne(ws, r,
                   [f"Total {communes[cc]}", Tc[0], Tc[1], Tc[2], Tc[3], pgc],
                   total=True)
        r += 1                                   # espace entre communes
    r = _source(ws, r)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Feuille 2 : Dénombrement par agent et par jour
# ---------------------------------------------------------------------------
def _feuille_agents(wb, conn, nom_district, menages):
    ws = wb.create_sheet("Dénombrement par agent-jour")
    ac = equipes.agents_et_chefs(conn)
    from collections import defaultdict

    # Dates observées (AAAAMMJJ valides) triées.
    dates = sorted({m["date"] for m in menages
                    if rapport_core._valid_date(m.get("date"))})

    def _jj(d):
        return f"{d[6:8]}/{d[4:6]}/{d[0:4]}" if len(d) == 8 else d

    # Regroupement : chef d'équipe -> (commune, agent) -> {date: n}. Le chef est
    # identifié par son NOM s'il est renseigné, sinon par son CODE (login_ce) —
    # ainsi « un tableau par chef » tient dès que le lien existe, même sans nom.
    cnt = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for m in menages:
        d = m.get("date")
        if not rapport_core._valid_date(d):
            continue
        commune = m.get("commune") or "(commune inconnue)"
        code = (m.get("agent") or "").strip()
        info = ac.get(code, {})
        agent_nom = info.get("nom") or code or "(agent inconnu)"
        chef = (info.get("chef_nom") or info.get("chef_login")
                or "(chef non affecté)")
        cnt[chef][(commune, agent_nom)][d] += 1

    # -- En-tête du document (bloc partagé, identique au « Rapport global ») --
    r = _entete_document(
        ws, nom_district,
        "Un tableau par chef d'équipe. Chaque valeur (colonnes de dates) "
        "= NOMBRE DE MÉNAGES DÉNOMBRÉS par l'agent ce jour-là.")
    r = _titre(ws, r, "Tableau — Dénombrement par agent et par jour "
               "(un tableau par chef d'équipe)")

    ncol = 2 + len(dates) + 1                     # Commune + Agent + dates + Total
    entete = ["Commune", "Agent"] + [_jj(d) for d in dates] + ["Total"]
    largeurs = [26, 28] + [11] * len(dates) + [10]

    # Un TABLEAU par chef d'équipe (même style que le « Rapport global »).
    def _tri_chef(k):
        return (1, k) if k.startswith("(chef") else (0, k)   # « non affecté » en dernier
    for chef in sorted(cnt, key=_tri_chef):
        r = _sous_titre(ws, r, f"Chef d'équipe : {chef}", ncol)
        r = _ecrire_entete(ws, r, entete, largeurs=largeurs)
        sous = cnt[chef]
        tot_col = [0] * len(dates)
        for (commune, agent_nom) in sorted(sous, key=lambda x: (x[0], x[1])):
            par_date = sous[(commune, agent_nom)]
            row = [commune, agent_nom]
            tot = 0
            for k, d in enumerate(dates):
                n = par_date.get(d, 0)
                row.append(n if n else "")
                tot += n
                tot_col[k] += n
            row.append(tot)
            r = _ligne(ws, r, row)
        r = _ligne(ws, r, [f"Total {chef}", ""]
                   + [g if g else "" for g in tot_col] + [sum(tot_col)], total=True)
        r = _source(ws, r)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Feuille 3 : BaseDenParAgent (table PLATE, une ligne par agent×zone)
# ---------------------------------------------------------------------------
def _feuille_base_agents(wb, conn, nom_district, menages):
    """Table « base » plate : une ligne par (commune, chef, agent, fokontany),
    colonnes = dates, cellule = nombre de ménages dénombrés. Pensée pour un
    tableau croisé dynamique / export brut."""
    ws = wb.create_sheet("BaseDenParAgent")
    ac = equipes.agents_et_chefs(conn)
    from collections import defaultdict

    dates = sorted({m["date"] for m in menages
                    if rapport_core._valid_date(m.get("date"))})

    def _jj(d):
        return f"{d[6:8]}/{d[4:6]}/{d[0:4]}" if len(d) == 8 else d

    # (commune, chef_nom, agent_nom, fokontany) -> {date: n}
    cnt = defaultdict(lambda: defaultdict(int))
    for m in menages:
        d = m.get("date")
        if not rapport_core._valid_date(d):
            continue
        commune = m.get("commune") or "(commune inconnue)"
        fkt = m.get("fkt") or m.get("fktcode") or ""
        code = (m.get("agent") or "").strip()
        info = ac.get(code, {})
        agent_nom = info.get("nom") or code or "(agent inconnu)"
        chef = (info.get("chef_nom") or info.get("chef_login")
                or "(chef non affecté)")
        cnt[(commune, chef, agent_nom, fkt)][d] += 1

    # -- En-tête du document (bloc partagé, identique au « Rapport global ») --
    r = _entete_document(
        ws, nom_district,
        "Une ligne par (commune, chef d'équipe, agent, fokontany). "
        "Chaque valeur de date = NOMBRE DE MÉNAGES DÉNOMBRÉS ce jour-là.")
    r = _titre(ws, r, "Tableau — Base dénombrement par agent et par jour")

    entete = (["Commune", "Chef d'équipe", "Agent", "Fokontany"]
              + [_jj(d) for d in dates] + ["Total"])
    largeurs = [24, 24, 26, 28] + [11] * len(dates) + [10]
    r = _ecrire_entete(ws, r, entete, largeurs=largeurs)

    grand = [0] * len(dates)
    for cle in sorted(cnt, key=lambda k: (k[0], k[1], k[2], k[3])):
        commune, chef, agent, fkt = cle
        par_date = cnt[cle]
        row = [commune, chef, agent, fkt]
        tot = 0
        for k, d in enumerate(dates):
            n = par_date.get(d, 0)
            row.append(n if n else "")
            tot += n
            grand[k] += n
        row.append(tot)
        r = _ligne(ws, r, row)
    r = _ligne(ws, r, ["TOTAL", "", "", ""]
               + [g if g else "" for g in grand] + [sum(grand)], total=True)
    r = _source(ws, r)
    # Fige les 4 colonnes de gauche + le bloc d'en-tête (jusqu'à la ligne d'en-tête
    # du tableau en ligne 8 : bloc document 1-4, titre 6, en-tête 8).
    ws.freeze_panes = "E9"


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------
def generer_classeur(conn, code_district, nom_district, communes_autorisees=None):
    """Renvoie le classeur Excel (openpyxl Workbook) du district."""
    menages = _menages_scope(conn, code_district, communes_autorisees)
    wb = Workbook()
    _feuille_global(wb, conn, code_district, nom_district, menages,
                    communes_autorisees)
    _feuille_agents(wb, conn, nom_district, menages)
    _feuille_base_agents(wb, conn, nom_district, menages)
    return wb


def generer_bytes(conn, code_district, nom_district, communes_autorisees=None):
    """Renvoie le classeur sérialisé (bytes) prêt à télécharger."""
    wb = generer_classeur(conn, code_district, nom_district, communes_autorisees)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
