# -*- coding: utf-8 -*-
"""
equipes.py — Base des Chefs d'Équipe (CE) et des Agents (AE).

Deux tables (créées au démarrage par `creer_tables`) :

    chef_equipe(
        login_ce       TEXT PRIMARY KEY,     -- identifiant du chef d'équipe
        nom_prenom_ce  TEXT NOT NULL)        -- nom et prénom du CE

    agent(
        login_ae       TEXT PRIMARY KEY,     -- identifiant de l'agent enquêteur
        nom_prenom_ae  TEXT NOT NULL,        -- nom et prénom de l'AE
        login_ce       TEXT REFERENCES chef_equipe(login_ce))   -- CE de rattachement

L'Expert « Traitement » remplit ces tables en TÉLÉVERSANT deux fichiers Excel (la
liste des CE et la liste des AE). La transcription est un UPSERT (comme maj_db) :
elle AJOUTE les nouveaux, MET À JOUR les modifiés, et NE SUPPRIME rien.

Note SQLite : les clés étrangères ne sont vérifiées que si `PRAGMA foreign_keys=ON`
(désactivé par défaut) ; l'intégrité (chaque agent rattaché à un CE existant) est de
toute façon garantie côté Python à l'import.

Colonnes Excel attendues (en-têtes, insensible à la casse ; alias tolérés) :
    CE     : login_ce, nom_prenom_ce   (alias nom : nom_prenom, « nom et prenom »)
    Agents : login_ae, nom_prenom_ae, login_ce   (alias nom : nom_prenom)
"""
import io

import admin           # style + _cellstr (réutilisés)
import db_source

ESC = admin.ESC
_cellstr = admin._cellstr


# ---------------------------------------------------------------------------
# Schéma
# ---------------------------------------------------------------------------
_DDL_CHEF = (
    'CREATE TABLE IF NOT EXISTS "chef_equipe" ('
    '"login_ce" TEXT PRIMARY KEY, "nom_prenom_ce" TEXT NOT NULL)')
_DDL_AGENT = (
    'CREATE TABLE IF NOT EXISTS "agent" ('
    '"login_ae" TEXT PRIMARY KEY, "nom_prenom_ae" TEXT NOT NULL, '
    '"login_ce" TEXT REFERENCES "chef_equipe" ("login_ce"))')


def creer_tables(conn) -> None:
    """Crée `chef_equipe` puis `agent` (FK) si absentes. Idempotent."""
    cur = conn.cursor()
    cur.execute(_DDL_CHEF)
    cur.execute(_DDL_AGENT)
    conn.commit()


# ---------------------------------------------------------------------------
# Accès / comptage
# ---------------------------------------------------------------------------
def compter(conn) -> dict:
    """{'chefs': n, 'agents': m} — nombre d'enregistrements de chaque table."""
    cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM "chef_equipe"')
    chefs = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM "agent"')
    agents = cur.fetchone()[0]
    return {"chefs": chefs, "agents": agents}


def _chef_existe(conn, login_ce) -> bool:
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(f'SELECT 1 FROM "chef_equipe" WHERE "login_ce"={ph}', (login_ce,))
    return cur.fetchone() is not None


# ---------------------------------------------------------------------------
# Lien avec les données de dénombrement (interview__diagnostics.responsible)
# ---------------------------------------------------------------------------
def synchroniser_agents(conn) -> int:
    """Complète `agent` à partir des codes agent présents dans les données de
    dénombrement (`interview__diagnostics.responsible`).

    Tout code présent dans le dénombrement mais ABSENT de la table `agent` y est
    inséré, avec le NOM = le CODE lui-même (et chef d'équipe inconnu). Ainsi chaque
    code agent des données a bien une ligne dans `agent` (intégrité de la clé
    étrangère garantie côté Python). Renvoie le nombre d'agents créés. Sans effet si
    la table de dénombrement n'existe pas encore (base neuve)."""
    creer_tables(conn)
    cur = conn.cursor()
    try:
        cur.execute('SELECT DISTINCT "responsible" FROM "interview__diagnostics" '
                    'WHERE "responsible" IS NOT NULL')
        codes = [r[0] for r in cur.fetchall()]
    except Exception:
        return 0                          # table de dénombrement absente
    ph = db_source._placeholder(conn)
    ajoutes = 0
    for code in codes:
        c = ("" if code is None else str(code)).strip()
        if not c:
            continue
        cur.execute(f'SELECT 1 FROM "agent" WHERE "login_ae"={ph}', (c,))
        if cur.fetchone() is None:
            cur.execute('INSERT INTO "agent" ("login_ae","nom_prenom_ae","login_ce") '
                        f'VALUES ({ph},{ph},NULL)', (c, c))   # nom = code
            ajoutes += 1
    conn.commit()
    return ajoutes


def noms_agents(conn) -> dict:
    """{code_agent: nom} pour les agents dont le nom est VRAIMENT renseigné
    (nom != code). Sert au rapport pour afficher le nom au lieu du code ; un code
    dont le nom n'a pas encore été renseigné (nom == code) n'y figure pas -> le
    rapport garde le code."""
    creer_tables(conn)
    cur = conn.cursor()
    cur.execute('SELECT "login_ae","nom_prenom_ae" FROM "agent"')
    out = {}
    for code, nom in cur.fetchall():
        c = ("" if code is None else str(code)).strip()
        n = ("" if nom is None else str(nom)).strip()
        if c and n and n != c:
            out[c] = n
    return out


def agents_et_chefs(conn) -> dict:
    """{ login_ae : {"nom": nom_agent, "chef_login": login_ce|"",
                     "chef_nom": nom_chef|""} } pour TOUS les agents connus.
    Sert à l'export (regroupement agent -> chef d'équipe)."""
    creer_tables(conn)
    cur = conn.cursor()
    cur.execute('SELECT a."login_ae", a."nom_prenom_ae", a."login_ce", '
                'c."nom_prenom_ce" FROM "agent" a '
                'LEFT JOIN "chef_equipe" c ON a."login_ce" = c."login_ce"')
    out = {}
    for lae, nae, lce, nce in cur.fetchall():
        code = ("" if lae is None else str(lae)).strip()
        if not code:
            continue
        out[code] = {
            "nom": ("" if nae is None else str(nae)).strip() or code,
            "chef_login": ("" if lce is None else str(lce)).strip(),
            "chef_nom": ("" if nce is None else str(nce)).strip(),
        }
    return out


# ---------------------------------------------------------------------------
# Lecture d'un Excel (openpyxl) -> lignes {colonne: valeur}
# ---------------------------------------------------------------------------
def _lire_xlsx(chemin_xlsx, alias):
    """Lit la 1re feuille. `alias` = {clef_logique: (noms d'en-tête acceptés)}.
    Renvoie (lignes, entetes_manquantes). Chaque ligne = {clef_logique: str}."""
    import openpyxl
    wb = openpyxl.load_workbook(chemin_xlsx, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    entetes = next(it, None)
    if not entetes:
        wb.close()
        return [], list(alias)          # fichier vide -> tout manque
    # index de colonne par en-tête normalisé (minuscule, espaces -> _)
    idx = {}
    for i, h in enumerate(entetes):
        if h is not None:
            cle = str(h).strip().lower().replace(" ", "_")
            idx.setdefault(cle, i)
    # résolution clef logique -> index via les alias
    col = {}
    manquantes = []
    for logique, noms in alias.items():
        trouve = next((idx[n] for n in noms if n in idx), None)
        if trouve is None:
            manquantes.append(logique)
        else:
            col[logique] = trouve
    if manquantes:
        wb.close()
        return [], manquantes
    lignes = []
    for r in it:
        if r is None or all(v is None for v in r):
            continue
        lignes.append({k: (_cellstr(r[i]) if i < len(r) else "")
                       for k, i in col.items()})
    wb.close()
    return lignes, []


_ALIAS_CHEF = {"login_ce": ("login_ce",),
               "nom_prenom_ce": ("nom_prenom_ce", "nom_prenom", "nom_et_prenom")}
_ALIAS_AGENT = {"login_ae": ("login_ae",),
                "nom_prenom_ae": ("nom_prenom_ae", "nom_prenom", "nom_et_prenom"),
                "login_ce": ("login_ce",)}


# ---------------------------------------------------------------------------
# Transcription (UPSERT) depuis les deux Excel
# ---------------------------------------------------------------------------
def _upsert_chef(conn, login_ce, nom):
    """Renvoie 'ajoute' | 'modifie' | 'inchange'."""
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(f'SELECT "nom_prenom_ce" FROM "chef_equipe" WHERE "login_ce"={ph}',
                (login_ce,))
    row = cur.fetchone()
    if row is None:
        cur.execute('INSERT INTO "chef_equipe" ("login_ce","nom_prenom_ce") '
                    f'VALUES ({ph},{ph})', (login_ce, nom))
        return "ajoute"
    if row[0] != nom:
        cur.execute(f'UPDATE "chef_equipe" SET "nom_prenom_ce"={ph} '
                    f'WHERE "login_ce"={ph}', (nom, login_ce))
        return "modifie"
    return "inchange"


def _upsert_agent(conn, login_ae, nom, login_ce):
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute('SELECT "nom_prenom_ae","login_ce" FROM "agent" '
                f'WHERE "login_ae"={ph}', (login_ae,))
    row = cur.fetchone()
    if row is None:
        cur.execute('INSERT INTO "agent" ("login_ae","nom_prenom_ae","login_ce") '
                    f'VALUES ({ph},{ph},{ph})', (login_ae, nom, login_ce))
        return "ajoute"
    if row[0] != nom or (row[1] or "") != (login_ce or ""):
        cur.execute(f'UPDATE "agent" SET "nom_prenom_ae"={ph},"login_ce"={ph} '
                    f'WHERE "login_ae"={ph}', (nom, login_ce, login_ae))
        return "modifie"
    return "inchange"


def _bilan_vide():
    return {"ajoutes": 0, "modifies": 0, "inchanges": 0}


def transcrire(conn, chemin_chef_xlsx, chemin_agent_xlsx):
    """Transcrit les deux Excel vers `chef_equipe` puis `agent` (upsert).

    Renvoie un dict :
        {"chefs": {ajoutes,modifies,inchanges}, "agents": {...},
         "erreurs": [(fichier, ligne, message), ...]}
    Lève ValueError si un fichier a des colonnes manquantes (rien n'est écrit).
    Les CE sont transcrits AVANT les agents (cible des clés étrangères)."""
    creer_tables(conn)
    chefs, manq_c = _lire_xlsx(chemin_chef_xlsx, _ALIAS_CHEF)
    if manq_c:
        raise ValueError("Fichier des Chefs d'Équipe — colonnes manquantes : "
                         + ", ".join(manq_c))
    agents, manq_a = _lire_xlsx(chemin_agent_xlsx, _ALIAS_AGENT)
    if manq_a:
        raise ValueError("Fichier des Agents — colonnes manquantes : "
                         + ", ".join(manq_a))

    res = {"chefs": _bilan_vide(), "agents": _bilan_vide(), "erreurs": []}
    # 1) Chefs d'équipe
    vus_ce = set()
    for n, ligne in enumerate(chefs, start=2):
        login_ce = ligne["login_ce"].strip()
        nom = ligne["nom_prenom_ce"].strip()
        if not login_ce:
            res["erreurs"].append(("Chefs d'Équipe", n, "login_ce vide"))
            continue
        if not nom:
            res["erreurs"].append(("Chefs d'Équipe", n,
                                   f"nom et prénom vides (login_ce={login_ce})"))
            continue
        res["chefs"][_upsert_chef(conn, login_ce, nom) + "s"] += 1
        vus_ce.add(login_ce)
    # 2) Agents (login_ce doit exister — dans le fichier CE ou déjà en base)
    for n, ligne in enumerate(agents, start=2):
        login_ae = ligne["login_ae"].strip()
        nom = ligne["nom_prenom_ae"].strip()
        login_ce = ligne["login_ce"].strip()
        if not login_ae:
            res["erreurs"].append(("Agents", n, "login_ae vide"))
            continue
        if not nom:
            res["erreurs"].append(("Agents", n,
                                   f"nom et prénom vides (login_ae={login_ae})"))
            continue
        if not login_ce:
            res["erreurs"].append(("Agents", n,
                                   f"login_ce (chef d'équipe) vide (agent {login_ae})"))
            continue
        if login_ce not in vus_ce and not _chef_existe(conn, login_ce):
            res["erreurs"].append(("Agents", n,
                                   f"chef d'équipe inconnu : {login_ce} "
                                   f"(agent {login_ae})"))
            continue
        res["agents"][_upsert_agent(conn, login_ae, nom, login_ce) + "s"] += 1
    conn.commit()
    return res


# ---------------------------------------------------------------------------
# Modèles Excel (canevas à remplir)
# ---------------------------------------------------------------------------
def modele_chef_xlsx() -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "chefs_equipe"
    ws.append(["login_ce", "nom_prenom_ce"])
    ws.append(["CE001", "RAKOTO Jean"])
    ws.append(["CE002", "RABE Marie"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def modele_agent_xlsx() -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "agents"
    ws.append(["login_ae", "nom_prenom_ae", "login_ce"])
    ws.append(["AE001", "RANDRIA Paul", "CE001"])
    ws.append(["AE002", "RASOA Hanta", "CE001"])
    ws.append(["AE003", "RAKOTOSON Luc", "CE002"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Pages (charte visuelle d'admin.py)
# ---------------------------------------------------------------------------
def _entete() -> str:
    return (f'<!doctype html><html lang="fr"><head><meta charset="utf-8">'
            f'<meta name="viewport" content="width=device-width,initial-scale=1">'
            f'<title>Traitement — RSU</title><style>{admin._STYLE}</style></head>'
            f'<body><div class="bar"><b>🗂 Espace Traitement</b>'
            f'<a href="/traitement">Accueil</a>'
            f'<a href="/choix">Tableau de bord</a>'
            f'<span class="sp"></span></div><div class="wrap">')


_CSS_CHOIX = """
<style>
.choix{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));
  gap:16px;margin-top:16px}
.ca{display:block;text-decoration:none;color:inherit;background:#fff;
  border:1.5px solid #e2e8f0;border-radius:14px;padding:20px 20px 18px;
  transition:.15s;position:relative}
.ca:hover{border-color:#2563eb;box-shadow:0 12px 26px rgba(37,99,235,.14);
  transform:translateY(-2px)}
.ca .ic{font-size:30px;line-height:1;margin-bottom:10px}
.ca .t{font-weight:700;font-size:16px;margin-bottom:6px}
.ca .d{font-size:13px;color:#64748b;line-height:1.5}
.ca .go{margin-top:12px;font-weight:700;color:#2563eb;font-size:13px}
</style>"""


def page_choix_traitement(district_txt) -> str:
    """Accueil de l'Expert Traitement : tableau de bord OU remplir CE/Agents."""
    h = [_entete(), _CSS_CHOIX, '<h1>Espace Traitement</h1>',
         f'<div class="note">District d’affectation : <b>{ESC(district_txt)}</b>.</div>',
         '<div class="choix">',
         '<a class="ca" href="/choix">'
         '<div class="ic">📊</div><div class="t">Tableau de bord (suivi)</div>'
         '<div class="d">Consulter le rapport de suivi du dénombrement pour votre '
         'district.</div><div class="go">Ouvrir →</div></a>',
         '<a class="ca" href="/traitement/equipes">'
         '<div class="ic">👥</div><div class="t">Remplir la base Chef d’Équipe et '
         'Agent</div>'
         '<div class="d">Téléverser les fichiers Excel des Chefs d’Équipe et des '
         'Agents, puis transcrire vers la base de données.</div>'
         '<div class="go">Ouvrir →</div></a>',
         '<a class="ca" href="/traitement/prechargement">'
         '<div class="ic">📦</div><div class="t">Générer la base de préchargement '
         '(VAD)</div>'
         '<div class="d">Produire la base de préchargement et le fichier de charge '
         'par agent (avec équilibrage optionnel) à partir du dénombrement transcrit '
         'de votre district.</div>'
         '<div class="go">Ouvrir →</div></a>',
         '</div>', '</div></body></html>']
    return "".join(h)


def _table_bilan(res) -> str:
    def ligne(nom, b):
        return (f'<tr><td>{nom}</td><td>+{b["ajoutes"]}</td>'
                f'<td>~{b["modifies"]}</td><td>={b["inchanges"]}</td></tr>')
    return ('<table><tr><th>Table</th><th>Ajoutés</th><th>Modifiés</th>'
            '<th>Inchangés</th></tr>'
            + ligne("Chefs d’Équipe", res["chefs"])
            + ligne("Agents", res["agents"]) + '</table>')


def page_equipes(conn, district_txt, resultat=None, message=None, erreur=None) -> str:
    """Formulaire de téléversement des deux Excel + bilan de la transcription."""
    n = compter(conn)
    h = [_entete(),
         '<p style="margin:0 0 6px"><a href="/traitement">← Accueil Traitement</a></p>',
         '<h1>Base des Chefs d’Équipe et des Agents</h1>',
         f'<div class="note">District d’affectation : <b>{ESC(district_txt)}</b>. '
         f'Actuellement en base : <b>{n["chefs"]}</b> chef(s) d’équipe et '
         f'<b>{n["agents"]}</b> agent(s).</div>']
    if message:
        h.append(f'<div class="msg">{ESC(message)}</div>')
    if erreur:
        h.append(f'<div class="err">{erreur}</div>')      # HTML autorisé (listes)

    if resultat:
        h.append('<h2>Bilan de la transcription</h2>')
        h.append(_table_bilan(resultat))
        if resultat.get("erreurs"):
            items = "".join(
                f'<li>{ESC(f)} — ligne {ln} : {ESC(m)}</li>'
                for f, ln, m in resultat["erreurs"])
            h.append('<div class="err"><b>Lignes ignorées :</b>'
                     f'<ul>{items}</ul></div>')
        h.append('<p><a href="/traitement/equipes">↻ Nouvelle transcription</a></p>')

    h.append('<h2>Téléverser les fichiers Excel</h2>')
    h.append(
        '<form method="post" action="/traitement/equipes" '
        'enctype="multipart/form-data" class="grid-form">'
        '<div><label>Fichier Excel — Liste des <b>Chefs d’Équipe</b> '
        '(colonnes : login_ce, nom_prenom_ce)</label>'
        '<input type="file" name="fichier_chef" accept=".xlsx" required></div>'
        '<div><label>Fichier Excel — Liste des <b>Agents</b> '
        '(colonnes : login_ae, nom_prenom_ae, login_ce)</label>'
        '<input type="file" name="fichier_agent" accept=".xlsx" required></div>'
        '<div style="align-self:end"><button>Transcrire vers la base</button></div>'
        '</form>')
    h.append('<div class="note">Les deux fichiers sont requis. La transcription '
             '<b>ajoute</b> les nouveaux, <b>met à jour</b> les modifiés et '
             '<b>ne supprime rien</b>. Chaque agent doit référencer un '
             '<b>login_ce</b> présent dans le fichier des Chefs d’Équipe ou déjà en '
             'base.<br>Modèles : '
             '<a href="/traitement/modele/chef.xlsx">⬇ Chefs d’Équipe</a> · '
             '<a href="/traitement/modele/agent.xlsx">⬇ Agents</a>.</div>')
    h.append('</div></body></html>')
    return "".join(h)
