# -*- coding: utf-8 -*-
"""
zones.py — Référentiel géographique (Province / Région / District / Commune /
Fokontany) pour l'application web RSU.

SOURCE : le fichier Excel « FKT_ampiasan_SS.xlsx » (découpage administratif
complet de Madagascar, ~20 000 fokontany).

SCHÉMA NORMALISÉ (3ᵉ forme normale) : une table par niveau administratif,
reliées par clés primaires / clés étrangères en relations un-à-plusieurs — au
lieu d'une seule grande table plate qui répéterait Province/Région/District sur
chaque ligne de fokontany (dépendances transitives = violation de la 3NF).

    province (code_province PK, nom)
    region   (code_region PK, nom, code_province -> province)
    district (code_district PK, nom, code_region  -> region)
    commune  (code_commune PK, nom, code_district -> district)
    fokontany(code_fokontany PK, nom, code_commune -> commune)

Une province a plusieurs régions ; une région plusieurs districts ; etc. Chaque
enfant porte la clé étrangère de son parent (le « côté un » de la relation).

Ce module :
    1. `charger_excel_vers_db(conn, xlsx)` (re)crée les 5 tables et les remplit.
    2. `assurer_zones(conn)` charge si les tables sont absentes.
    3. `arbre_geo(conn)` renvoie de quoi construire les listes déroulantes
       dépendantes Province -> Région -> District de la page de sélection.
    4. `libelles_district(conn, code_district)` -> (province, region, district).

PORTABLE SQLite <-> PostgreSQL, comme db_source : tout passe par la DB-API 2.0.
Le MÊME code remplit une base SQLite (défaut, zéro installation) ou PostgreSQL
(production, quand RSU_DB_URL est défini). Voir db_source.connect().

    # Charger dans SQLite (défaut) :
    python zones.py

    # Charger dans PostgreSQL :
    set RSU_DB_URL=postgresql://user:mdp@localhost:5432/rsu
    python zones.py
"""

import os

import db_source

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_DEFAUT = os.path.join(BASE, "FKT_ampiasan_SS.xlsx")
# Excel des ménages estimés 2025 par commune (issu de FKT_ampiasan_SS croisé avec
# la projection population INSTAT) : sert à remplir commune."nombreMenage".
MENAGES_XLSX = os.path.join(BASE, "MENAGES_PAR_COMMUNE_2025.xlsx")

# Colonnes attendues dans l'Excel (par leur nom d'en-tête).
COLS_EXCEL = [
    "code_province", "province",
    "code_region", "region",
    "code_district", "district",
    "code_commune", "commune",
    "code_fokontany", "fokontany",
]

# Définition des 5 tables normalisées, du parent vers l'enfant.
# (table, colonne_code_PK, colonne_code_parent_FK, table_parent)
NIVEAUX = [
    ("province", "code_province", None, None),
    ("region", "code_region", "code_province", "province"),
    ("district", "code_district", "code_region", "region"),
    ("commune", "code_commune", "code_district", "district"),
    ("fokontany", "code_fokontany", "code_commune", "commune"),
]


# ---------------------------------------------------------------------------
# Lecture de l'Excel
# ---------------------------------------------------------------------------
def _lire_excel(xlsx_path: str):
    """Renvoie (liste de lignes, index par nom de colonne). Une ligne = tuple
    des 10 valeurs, dans l'ordre de COLS_EXCEL."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    entetes = next(it, None)
    if not entetes:
        raise ValueError("Fichier Excel vide.")
    noms = [(str(h).strip().lower() if h is not None else "") for h in entetes]
    idx = {}
    for cle in COLS_EXCEL:
        if cle not in noms:
            raise ValueError(
                f"Colonne « {cle} » absente de l'Excel. En-têtes : {noms}")
        idx[cle] = noms.index(cle)

    lignes = []
    for r in it:
        if r is None or all(v is None for v in r):
            continue
        lignes.append({cle: r[idx[cle]] for cle in COLS_EXCEL})
    wb.close()
    return lignes


# ---------------------------------------------------------------------------
# ETL : Excel -> 5 tables normalisées
# ---------------------------------------------------------------------------
def _extraire_niveaux(lignes):
    """Déduplique chaque niveau par son code. Renvoie un dict :
    { "province": {code: (nom,)}, "region": {code: (nom, code_parent)}, ... }.
    """
    donnees = {t: {} for t, _pk, _fk, _p in NIVEAUX}
    for L in lignes:
        # province (pas de parent)
        cp = L["code_province"]
        if cp is not None:
            donnees["province"].setdefault(cp, (L["province"],))
        # region -> province
        cr = L["code_region"]
        if cr is not None:
            donnees["region"].setdefault(cr, (L["region"], cp))
        # district -> region
        cd = L["code_district"]
        if cd is not None:
            donnees["district"].setdefault(cd, (L["district"], cr))
        # commune -> district
        cc = L["code_commune"]
        if cc is not None:
            donnees["commune"].setdefault(cc, (L["commune"], cd))
        # fokontany -> commune
        cf = L["code_fokontany"]
        if cf is not None:
            donnees["fokontany"].setdefault(cf, (L["fokontany"], cc))
    return donnees


def charger_excel_vers_db(conn, xlsx_path: str = EXCEL_DEFAUT, log=print):
    """(Re)crée les 5 tables normalisées et les remplit depuis l'Excel."""
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"Fichier Excel introuvable : {xlsx_path}")
    log(f"Lecture de l'Excel : {os.path.basename(xlsx_path)}…")
    lignes = _lire_excel(xlsx_path)
    donnees = _extraire_niveaux(lignes)

    ph = db_source._placeholder(conn)
    cur = conn.cursor()

    # Nettoyage : ancienne table plate + tables normalisées (enfants -> parents).
    cur.execute('DROP TABLE IF EXISTS "zones"')  # ancienne version dénormalisée
    for table, _pk, _fk, _parent in reversed(NIVEAUX):
        cur.execute(f'DROP TABLE IF EXISTS "{table}"')

    # Création : parents -> enfants (pour que les clés étrangères existent).
    for table, pk, fk, parent in NIVEAUX:
        cols = [f'"{pk}" BIGINT PRIMARY KEY', '"nom" TEXT NOT NULL']
        if fk:
            cols.append(f'"{fk}" BIGINT NOT NULL '
                        f'REFERENCES "{parent}" ("{fk}")')
        # La commune porte en plus le nombre de ménages estimés (nullable ;
        # rempli ensuite par charger_menages depuis MENAGES_XLSX).
        if table == "commune":
            cols.append('"nombreMenage" BIGINT')
        cur.execute(f'CREATE TABLE "{table}" ({", ".join(cols)})')

    # Insertion : parents -> enfants.
    comptes = {}
    for table, pk, fk, _parent in NIVEAUX:
        table_data = donnees[table]
        if fk:
            colonnes = f'"{pk}","nom","{fk}"'
            marks = f"{ph},{ph},{ph}"
            valeurs = [(code, nom, parent_code)
                       for code, (nom, parent_code) in table_data.items()]
        else:
            colonnes = f'"{pk}","nom"'
            marks = f"{ph},{ph}"
            valeurs = [(code, nom) for code, (nom,) in table_data.items()]
        cur.executemany(
            f'INSERT INTO "{table}" ({colonnes}) VALUES ({marks})', valeurs)
        comptes[table] = len(valeurs)

        # Index sur la clé étrangère (jointures / listes dépendantes).
        if fk:
            try:
                cur.execute(
                    f'CREATE INDEX "ix_{table}_{fk}" ON "{table}" ("{fk}")')
            except Exception:
                pass

    conn.commit()
    log("Tables normalisées chargées (commit) : " +
        ", ".join(f"{t}={comptes[t]}" for t, _pk, _fk, _p in NIVEAUX))

    # Remplissage du nombre de ménages par commune si l'Excel dérivé est présent.
    if os.path.isfile(MENAGES_XLSX):
        try:
            n = charger_menages(conn, MENAGES_XLSX, log=log)
            comptes["menages"] = n
        except Exception as e:  # ne bloque pas le chargement du référentiel
            log(f"(nombreMenage non rempli : {e})")
    return comptes


# ---------------------------------------------------------------------------
# Nombre de ménages par commune (colonne commune."nombreMenage")
# ---------------------------------------------------------------------------
def _colonnes_table(conn, table):
    """Ensemble des noms de colonnes d'une table (SQLite ou PostgreSQL)."""
    cur = conn.cursor()
    mod = type(conn).__module__.split(".")[0]
    if mod.startswith("psycopg"):
        cur.execute("SELECT column_name FROM information_schema.columns "
                    "WHERE table_name = %s", (table,))
        return {r[0] for r in cur.fetchall()}
    cur.execute(f'PRAGMA table_info("{table}")')
    return {r[1] for r in cur.fetchall()}


def assurer_colonne_menages(conn):
    """Ajoute la colonne commune.\"nombreMenage\" si elle manque (idempotent)."""
    if "nombreMenage" in _colonnes_table(conn, "commune"):
        return False
    conn.cursor().execute('ALTER TABLE "commune" ADD COLUMN "nombreMenage" BIGINT')
    conn.commit()
    return True


def _lire_menages_excel(xlsx_path):
    """Lit MENAGES_XLSX -> {code_commune(int): nombre_menage(int)}.
    Repère les colonnes par en-tête : « code_commune » et « menages_estimes_2025 »."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["menages_par_commune"] if "menages_par_commune" in wb.sheetnames \
        else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    entetes = [(str(h).strip().lower() if h is not None else "") for h in next(it)]
    try:
        i_code = entetes.index("code_commune")
        i_men = entetes.index("menages_estimes_2025")
    except ValueError:
        wb.close()
        raise ValueError(
            f"Colonnes « code_commune » / « menages_estimes_2025 » absentes. "
            f"En-têtes : {entetes}")
    out = {}
    for r in it:
        if r is None or r[i_code] is None:
            continue
        val = r[i_men]
        out[int(r[i_code])] = int(round(val)) if isinstance(val, (int, float)) else None
    wb.close()
    return out


def attendus_communes(conn, ccodes):
    """{ str(code_commune) : {"nom": nom, "attendu": nombreMenage(int|None)} }
    pour les codes commune donnés (itérable de codes str/int). Sert au calcul de
    couverture (dénombrement réalisé vs projection ménages)."""
    codes = sorted({int(c) for c in ccodes if str(c).strip().isdigit()})
    if not codes:
        return {}
    if "nombreMenage" not in _colonnes_table(conn, "commune"):
        assurer_colonne_menages(conn)
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    marks = ",".join([ph] * len(codes))
    cur.execute(f'SELECT code_commune, nom, "nombreMenage" FROM "commune" '
                f'WHERE code_commune IN ({marks})', codes)
    return {str(code): {"nom": nom, "attendu": men}
            for code, nom, men in cur.fetchall()}


def charger_menages(conn, xlsx_path: str = MENAGES_XLSX, log=print):
    """Remplit commune.\"nombreMenage\" depuis l'Excel des ménages (par code_commune).
    Ajoute la colonne au besoin. Renvoie le nombre de communes mises à jour."""
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"Excel des ménages introuvable : {xlsx_path}")
    assurer_colonne_menages(conn)
    donnees = _lire_menages_excel(xlsx_path)
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    maj = 0
    for code, men in donnees.items():
        cur.execute(f'UPDATE "commune" SET "nombreMenage" = {ph} '
                    f'WHERE code_commune = {ph}', (men, code))
        maj += cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0
    conn.commit()
    log(f"nombreMenage rempli : {maj} communes mises à jour "
        f"(sur {len(donnees)} lignes Excel).")
    return maj


def zones_existe(conn) -> bool:
    """Vrai si le référentiel normalisé est présent et peuplé."""
    cur = conn.cursor()
    try:
        cur.execute('SELECT COUNT(*) FROM "province"')
        if cur.fetchone()[0] == 0:
            return False
        cur.execute('SELECT COUNT(*) FROM "fokontany"')
        return cur.fetchone()[0] > 0
    except Exception:
        return False


def assurer_zones(conn, xlsx_path: str = EXCEL_DEFAUT, log=print):
    """Charge le référentiel depuis l'Excel s'il est absent/vide."""
    if zones_existe(conn):
        return False
    charger_excel_vers_db(conn, xlsx_path, log)
    return True


# ---------------------------------------------------------------------------
# Arbre géographique pour les listes déroulantes dépendantes
# ---------------------------------------------------------------------------
def arbre_geo(conn):
    """Renvoie un dict prêt pour un cascade Province -> Région -> District :

        {
          "provinces": [{"c": 3, "n": "Fianarantsoa"}, ...],       (triées)
          "regions":   {"3": [{"c": 36, "n": "VATOVAVY"}, ...]},   (par code_province)
          "districts": {"36": [{"c": 3602, "n": "MANANJARY"}, ...]}(par code_region)
        }
    """
    cur = conn.cursor()

    cur.execute('SELECT code_province, nom FROM "province" ORDER BY nom')
    provinces = [{"c": c, "n": n} for c, n in cur.fetchall()]

    cur.execute('SELECT code_province, code_region, nom FROM "region" '
                'ORDER BY nom')
    regions = {}
    for cp, cr, n in cur.fetchall():
        regions.setdefault(str(cp), []).append({"c": cr, "n": n})

    cur.execute('SELECT code_region, code_district, nom FROM "district" '
                'ORDER BY nom')
    districts = {}
    for cr, cd, n in cur.fetchall():
        districts.setdefault(str(cr), []).append({"c": cd, "n": n})

    return {"provinces": provinces, "regions": regions, "districts": districts}


def libelles_district(conn, code_district):
    """Renvoie (province, region, district) pour un code_district, ou None.

    Jointure district -> region -> province (les relations normalisées)."""
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(
        f'SELECT p.nom, r.nom, d.nom '
        f'FROM "district" d '
        f'JOIN "region" r ON d.code_region = r.code_region '
        f'JOIN "province" p ON r.code_province = p.code_province '
        f'WHERE d.code_district = {ph} LIMIT 1', (int(code_district),))
    return cur.fetchone()


def codes_fokontany_district(conn, code_district):
    """Renvoie la liste des codes fokontany (str, 8 chiffres) d'un district.

    Sert à afficher les contours (carte) de tout le district même quand il n'a
    aucun ménage dénombré — voir rapport_core.generer_rapport(codes_geo=...)."""
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(
        f'SELECT f.code_fokontany FROM "fokontany" f '
        f'JOIN "commune" cm ON f.code_commune = cm.code_commune '
        f'WHERE cm.code_district = {ph}', (int(code_district),))
    return [str(r[0]) for r in cur.fetchall()]


def reference_district(conn, code_district):
    """Liste de référence COMPLÈTE des unités d'un district (depuis `zones`).

    Renvoie [{"fkt": "<code8>", "label": "<nom fokontany>",
              "commune": "<nom commune>", "ccode": "<code6>"}, ...] pour TOUS les
    fokontany du district — y compris ceux sans données de dénombrement. Le
    gabarit fusionne cette liste pour afficher toutes les communes/fokontany et
    repérer les zones oubliées (voir rapport_core.generer_rapport(zones_ref=...))."""
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(
        f'SELECT f.code_fokontany, f.nom, cm.code_commune, cm.nom '
        f'FROM "fokontany" f '
        f'JOIN "commune" cm ON f.code_commune = cm.code_commune '
        f'WHERE cm.code_district = {ph} '
        f'ORDER BY cm.nom, f.nom', (int(code_district),))
    return [{"fkt": str(fc), "label": fn, "commune": cn, "ccode": str(cc)}
            for fc, fn, cc, cn in cur.fetchall()]


def tous_districts(conn):
    """[{code, nom, region, province}] de TOUS les districts, triés par nom.

    Sert à choisir le district d'affectation d'un utilisateur (utilisateurs.py)."""
    cur = conn.cursor()
    cur.execute('SELECT d.code_district, d.nom, r.nom, p.nom '
                'FROM "district" d '
                'JOIN "region" r ON d.code_region = r.code_region '
                'JOIN "province" p ON r.code_province = p.code_province '
                'ORDER BY d.nom')
    return [{"code": str(cd), "nom": nd, "region": nr, "province": pv}
            for cd, nd, nr, pv in cur.fetchall()]


def communes_district(conn, code_district):
    """[{code, nom}] des communes d'un district, triées par nom."""
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(f'SELECT code_commune, nom FROM "commune" '
                f'WHERE code_district = {ph} ORDER BY nom', (int(code_district),))
    return [{"code": str(cc), "nom": nc} for cc, nc in cur.fetchall()]


def communes_par_district(conn):
    """{ "<code_district>": [{"c": code_commune, "n": nom}, ...] } — pour un cascade
    District -> Commune côté formulaire (même forme {c,n} que arbre_geo)."""
    cur = conn.cursor()
    cur.execute('SELECT code_district, code_commune, nom FROM "commune" '
                'ORDER BY code_district, nom')
    out = {}
    for cd, cc, n in cur.fetchall():
        out.setdefault(str(cd), []).append({"c": cc, "n": n})
    return out


def commune_dans_district(conn, code_commune, code_district) -> bool:
    """Vrai si la commune appartient bien au district (intégrité d'affectation)."""
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(f'SELECT 1 FROM "commune" WHERE code_commune = {ph} '
                f'AND code_district = {ph}', (int(code_commune), int(code_district)))
    return cur.fetchone() is not None


def district_de_commune(conn, code_commune):
    """Code district (int) d'une commune, ou None si la commune est inconnue.
    Sert à déduire le district d'affectation d'un rôle « district + communes »."""
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(f'SELECT code_district FROM "commune" WHERE code_commune = {ph} '
                'LIMIT 1', (int(code_commune),))
    r = cur.fetchone()
    return int(r[0]) if r and r[0] is not None else None


def libelle_commune(conn, code_commune):
    """Nom de la commune (ou None si inconnue)."""
    ph = db_source._placeholder(conn)
    cur = conn.cursor()
    cur.execute(f'SELECT nom FROM "commune" WHERE code_commune = {ph} LIMIT 1',
                (int(code_commune),))
    r = cur.fetchone()
    return r[0] if r else None


# ---------------------------------------------------------------------------
# Exécution directe : charge la base pointée par RSU_DB_URL (ou SQLite).
# ---------------------------------------------------------------------------
def main(argv=None):
    import sys
    argv = sys.argv[1:] if argv is None else argv
    conn = db_source.connect()
    try:
        if argv and argv[0] == "menages":
            # Remplir uniquement commune."nombreMenage" (sans recharger les zones).
            xlsx = argv[1] if len(argv) > 1 else MENAGES_XLSX
            charger_menages(conn, xlsx)
            return
        charger_excel_vers_db(conn)
        arbre = arbre_geo(conn)
        print(f"Vérification : {len(arbre['provinces'])} provinces, "
              f"{sum(len(v) for v in arbre['regions'].values())} régions, "
              f"{sum(len(v) for v in arbre['districts'].values())} districts.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
