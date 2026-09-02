# -*- coding: utf-8 -*-
"""
admin.py — Page d'administration (réservée au rôle Admin).

Regroupe la LOGIQUE et le RENDU de l'espace admin, pour garder serveur_app léger :
    - stats(conn)            : comptes par rôle, actifs/désactivés
    - couverture(conn)       : districts/communes sans responsable affecté
    - importer_excel(...)    : création en masse depuis un .xlsx (mdp hachés)
    - modele_xlsx()          : modèle Excel d'import (octets)
    - export_journal_csv / export_utilisateurs_csv
    - page_admin(...) / page_admin_utilisateurs(...) : HTML

Réutilise utilisateurs (comptes), journal (audit), zones (référentiel).
"""
import csv
import html
import io
import json
import urllib.parse

import journal
import utilisateurs
import zones

ESC = html.escape


# ---------------------------------------------------------------------------
# Statistiques et couverture
# ---------------------------------------------------------------------------
def stats(conn) -> dict:
    comptes = utilisateurs.lister(conn)
    par_role = {}
    for u in comptes:
        par_role[u["responsabilite"]] = par_role.get(u["responsabilite"], 0) + 1
    actifs = sum(1 for u in comptes if u["actif"])
    return {"total": len(comptes),
            "par_role": {r: par_role.get(r, 0) for r in utilisateurs.RESPONSABILITES},
            "actifs": actifs, "inactifs": len(comptes) - actifs}


def couverture(conn) -> list:
    """Pour chaque district ayant au moins un utilisateur affecté : nb de responsables
    et communes SANS superviseur. Met en évidence les zones oubliées."""
    comptes = utilisateurs.lister(conn)
    resp_par_district = {}          # code_district -> nb responsables
    superv_communes = {}            # code_district -> set(codes communes supervisées)
    for u in comptes:
        # Districts couverts : celui d'affectation directe ET ceux d'un rôle
        # multi-district (liaison responsable_district).
        dcodes = [str(d) for d in (u.get("districts_affectation") or [])]
        if u["district_affectation"]:
            dcodes.append(str(u["district_affectation"]))
        for d in dcodes:
            resp_par_district[d] = resp_par_district.get(d, 0) + 1
        # Communes supervisées (rôles « district + communes »).
        if (u["responsabilite"] in utilisateurs._ROLES_DISTRICT_COMMUNES
                and u["district_affectation"]):
            superv_communes.setdefault(str(u["district_affectation"]), set()).update(
                str(c) for c in u["communes_affectation"])
    lignes = []
    for d in sorted(resp_par_district):
        lib = zones.libelles_district(conn, d)
        communes = zones.communes_district(conn, d)
        couvertes = superv_communes.get(d, set())
        sans = [c for c in communes if c["code"] not in couvertes]
        lignes.append({
            "code": d, "nom": lib[2] if lib else d,
            "nb_responsables": resp_par_district[d],
            "nb_communes": len(communes),
            "communes_sans_superviseur": sans,
        })
    return lignes


# ---------------------------------------------------------------------------
# Import / export
# ---------------------------------------------------------------------------
def _cellstr(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def importer_excel(conn, chemin_xlsx):
    """Crée des utilisateurs depuis un .xlsx. Colonnes attendues (en-têtes, insensible
    à la casse) : nom_prenom, responsabilite, login, mot_de_passe, district, communes.
    Renvoie (nb_ajoutes, [(ligne, message_erreur), ...])."""
    import openpyxl
    wb = openpyxl.load_workbook(chemin_xlsx, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    entetes = next(it, None)
    if not entetes:
        wb.close()
        return 0, [(1, "Fichier vide.")]
    idx = {}
    for i, h in enumerate(entetes):
        if h is not None:
            idx[str(h).strip().lower()] = i

    def val(col, r):
        i = idx.get(col)
        return _cellstr(r[i]) if i is not None and i < len(r) else ""

    manquantes = [c for c in ("nom_prenom", "responsabilite", "login", "mot_de_passe")
                  if c not in idx]
    if manquantes:
        wb.close()
        return 0, [(1, "Colonnes manquantes : " + ", ".join(manquantes))]

    ajoutes, erreurs = 0, []
    for n, r in enumerate(it, start=2):
        if r is None or all(v is None for v in r):
            continue
        try:
            utilisateurs.ajouter(
                conn, val("login", r), val("nom_prenom", r),
                val("responsabilite", r), val("mot_de_passe", r),
                district_affectation=val("district", r) or None,
                commune_affectation=val("communes", r) or None,
                districts_affectation=val("districts", r) or None,
                telephone=val("telephone", r) or None,
                cin=val("cin", r) or None,
                email=val("email", r) or None,
                numero_orange_float=val("numero_orange_float", r) or None,
                sexe=val("sexe", r) or None)
            ajoutes += 1
        except Exception as e:                      # ValueError de validation, etc.
            erreurs.append((n, str(e)))
    wb.close()
    conn.commit()
    return ajoutes, erreurs


def modele_xlsx() -> bytes:
    """Modèle Excel d'import (en-têtes + exemples)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "utilisateurs"
    # Ordre des colonnes = ordre du formulaire (nom, rôle, coordonnées, identifiants,
    # affectation). L'import lit par nom d'en-tête : l'ordre est indicatif.
    ws.append(["nom_prenom", "responsabilite", "telephone", "cin", "email",
               "numero_orange_float", "sexe", "login", "mot_de_passe", "district",
               "communes", "districts"])
    ws.append(["RAKOTO Jean", "Superviseur Technique", "0340000001", "101012345678",
               "rakoto@example.mg", "0320000010", "Masculin", "rakoto", "MotDePasse#1",
               "4405", "440501,440502", ""])
    ws.append(["RABE Marie", "Expert survey", "0320000002", "101087654321",
               "rabe@example.mg", "0320000011", "Féminin", "rabe", "MotDePasse#2",
               "4405", "", ""])
    ws.append(["Le Coordonnateur", "Coordonnateur Nationale", "", "", "", "", "",
               "coordnat", "MotDePasse#3", "", "", ""])
    ws.append(["RANDRIA Paul", "Coordonnateur régionale", "", "", "", "", "Masculin",
               "reg", "MotDePasse#4", "", "", "3301,3302,3303"])
    ws.append(["Log InterCommunale", "Logistique Inter-Communale", "", "", "", "", "",
               "logc", "MotDePasse#5", "4405", "440501", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_journal_csv(conn) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["login", "nom_prenom", "responsabilite", "ip",
                "connexion_le", "deconnexion_le", "duree", "statut"])
    for r in journal.journal(conn, limite=1000000):
        w.writerow([r["login"], r["nom_prenom"], r["responsabilite"], r["ip"],
                    r["connexion_le"], r["deconnexion_le"], r["duree"], r["statut"]])
    return buf.getvalue().encode("utf-8-sig")       # BOM -> Excel ouvre en UTF-8


def export_utilisateurs_csv(conn) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["login", "nom_prenom", "responsabilite", "affectation",
                "telephone", "cin", "email", "numero_orange_float", "sexe",
                "actif", "cree_le"])
    for u in utilisateurs.lister(conn):
        aff = utilisateurs.affectation_texte(conn, u["district_affectation"],
                                             u["communes_affectation"],
                                             u.get("districts_affectation"))
        w.writerow([u["login"], u["nom_prenom"], u["responsabilite"], aff,
                    u.get("telephone") or "", u.get("cin") or "",
                    u.get("email") or "", u.get("numero_orange_float") or "",
                    u.get("sexe") or "", "oui" if u["actif"] else "non", u["cree_le"]])
    return buf.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Rendu HTML
# ---------------------------------------------------------------------------
_STYLE = """
:root{--rsu-fluid:1;font-size:clamp(12px, 0.22vw + 10.2px, 14px);}
*{box-sizing:border-box} body{margin:0;font-family:'Segoe UI',system-ui,sans-serif;
  background:#f0f2f7;color:#1e293b}
.bar{background:#0f172a;color:#fff;padding:0.75rem 1.375rem;display:flex;align-items:center;
  gap:1.125rem;flex-wrap:wrap}
.bar b{font-size:0.9375rem} .bar a{color:#cbd5e1;text-decoration:none;font-size:0.8125rem}
.bar a:hover{color:#fff} .bar .sp{flex:1}
.wrap{max-width:67.5rem;margin:1.375rem auto;padding:0 1rem}
h2{font-size:1rem;margin:1.625rem 0 0.75rem} h1{font-size:1.25rem;margin:0 0 0.25rem}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(9.375rem,1fr));gap:0.75rem}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:0.75rem;padding:0.875rem 1rem}
.card .n{font-size:1.625rem;font-weight:800} .card .l{font-size:0.6875rem;color:#64748b;
  text-transform:uppercase;letter-spacing:.5px}
table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #e2e8f0;
  border-radius:0.75rem;overflow:hidden;font-size:0.8125rem}
th,td{padding:0.5rem 0.625rem;text-align:left;border-bottom:1px solid #eef1f6}
th{background:#f8fafc;font-size:0.6875rem;text-transform:uppercase;color:#64748b}
tr:last-child td{border-bottom:none}
.pill{padding:2px 0.5rem;border-radius:1.25rem;font-size:0.6875rem;font-weight:600}
.ok{background:#dcfce7;color:#166534}.ko{background:#fee2e2;color:#991b1b}
.warn{background:#fef9c3;color:#854d0e}
form.inline{display:inline}
input,select{padding:0.4375rem 0.5625rem;border:1px solid #cbd5e1;border-radius:0.5rem;font-size:0.8125rem;
  font-family:inherit}
button{padding:0.4375rem 0.75rem;border:none;border-radius:0.5rem;background:#2563eb;color:#fff;
  font-size:0.75rem;font-weight:600;cursor:pointer}
button.sec{background:#e2e8f0;color:#1e293b} button.danger{background:#ef4444}
.grid-form{display:grid;grid-template-columns:repeat(auto-fit,minmax(12.5rem,1fr));
  gap:0.625rem;background:#fff;border:1px solid #e2e8f0;border-radius:0.75rem;padding:1rem}
.note{background:#eff6ff;border:1px solid #bfdbfe;border-radius:0.625rem;padding:0.625rem 0.875rem;
  font-size:0.8125rem;margin:0.625rem 0}
.msg{background:#dcfce7;border:1px solid #86efac;border-radius:0.625rem;padding:0.625rem 0.875rem;
  font-size:0.8125rem;margin:0.625rem 0}
.err{background:#fee2e2;border:1px solid #fecaca;border-radius:0.625rem;padding:0.625rem 0.875rem;
  font-size:0.8125rem;margin:0.625rem 0}
label{font-size:0.75rem;color:#475569;display:block;margin-bottom:0.1875rem}
small{color:#64748b}
"""


def _entete(actif="") -> str:
    lien = lambda h, t, k: (f'<a href="{h}" style="color:#fff;font-weight:600">{t}</a>'
                            if k == actif else f'<a href="{h}">{t}</a>')
    return (f'<!doctype html><html lang="fr"><head><meta charset="utf-8">'
            f'<meta name="viewport" content="width=device-width,initial-scale=1">'
            f'<title>Administration RSU</title><style>{_STYLE}</style></head><body>'
            f'<div class="bar"><b>⚙ Administration RSU</b>'
            f'{lien("/admin","Tableau de bord","accueil")}'
            f'{lien("/admin/utilisateurs","Utilisateurs","users")}'
            f'{lien("/admin/utilisateurs/ajouter","Ajouter","users_add")}'
            f'{lien("/admin/journal","Journal","journal")}'
            f'<a href="/">Application</a>'
            f'<span class="sp"></span></div>'  # deconnexion = bandeau (haut droite)
            f'<div class="wrap">')


def _pied() -> str:
    return "</div></body></html>"


def _table_transcriptions(rows) -> str:
    """Historique des transcriptions (date+heure, personne, district, issue)."""
    lignes = "".join(
        f'<tr><td>{ESC(x["quand_court"])}</td>'
        f'<td>{ESC(x["nom_prenom"] or x["login"] or "—")}</td>'
        f'<td>{ESC(x["district"])}</td>'
        f'<td>{ESC(x["evenement"])}</td>'
        f'<td><span class="pill {"ok" if x["statut"] == "Réussi" else "ko"}">'
        f'{ESC(x["statut"])}</span></td>'
        f'<td>{ESC(x["detail"])}</td></tr>' for x in rows) \
        or ('<tr><td colspan="6"><small>Aucune transcription enregistrée.'
            '</small></td></tr>')
    return ('<table><tr><th>Date &amp; heure</th><th>Personne</th><th>District</th>'
            '<th>Opération</th><th>Statut</th><th>Détail</th></tr>'
            + lignes + '</table>')


def page_admin(conn, utilisateur, nb_connectes=0) -> str:
    st = stats(conn)
    cartes = (f'<div class="card"><div class="n">{st["total"]}</div>'
              f'<div class="l">Utilisateurs</div></div>'
              f'<div class="card"><div class="n">{st["actifs"]}</div>'
              f'<div class="l">Actifs</div></div>'
              f'<div class="card"><div class="n">{st["inactifs"]}</div>'
              f'<div class="l">Désactivés</div></div>'
              f'<div class="card"><div class="n">{nb_connectes}</div>'
              f'<div class="l">Connectés</div></div>')
    roles = "".join(f'<div class="card"><div class="n">{n}</div>'
                    f'<div class="l">{ESC(r)}</div></div>'
                    for r, n in st["par_role"].items())

    # Journal récent (10 dernières sessions)
    jr = journal.journal(conn, limite=10)
    lignes_j = "".join(
        f'<tr><td>{ESC(x["nom_prenom"] or x["login"] or "—")}</td>'
        f'<td>{ESC(x["responsabilite"] or "—")}</td>'
        f'<td>{ESC(x["connexion_le"] or "—")}</td><td>{ESC(x["duree"])}</td>'
        f'<td><span class="pill {"ok" if x["statut"]=="terminée" else "warn"}">'
        f'{x["statut"]}</span></td></tr>' for x in jr) \
        or '<tr><td colspan="5"><small>Aucune connexion enregistrée.</small></td></tr>'

    # Tentatives échouées (5 dernières)
    te = journal.tentatives(conn, limite=5)
    lignes_t = "".join(
        f'<tr><td>{ESC(x["login_saisi"] or "—")}</td><td>{ESC(x["quand"])}</td>'
        f'<td>{ESC(x["motif"])}</td></tr>' for x in te) \
        or '<tr><td colspan="3"><small>Aucune tentative échouée.</small></td></tr>'

    # Transcriptions récentes (téléversements + transcriptions, réussis ou non)
    tr = journal.transcriptions(conn, limite=15)

    # Couverture
    cov = couverture(conn)
    if cov:
        lignes_c = "".join(
            f'<tr><td>{ESC(c["nom"])} ({c["code"]})</td>'
            f'<td>{c["nb_responsables"]}</td>'
            f'<td>{len(c["communes_sans_superviseur"])} / {c["nb_communes"]}'
            + (': ' + ", ".join(ESC(x["nom"]) for x in c["communes_sans_superviseur"][:8])
               + ('…' if len(c["communes_sans_superviseur"]) > 8 else '')
               if c["communes_sans_superviseur"] else '')
            + '</td></tr>' for c in cov)
    else:
        lignes_c = ('<tr><td colspan="3"><small>Aucun district avec responsable '
                    'affecté pour l’instant.</small></td></tr>')

    return (_entete("accueil")
            + f'<h1>Bonjour {ESC(utilisateur.get("nom_prenom") or utilisateur.get("login"))}</h1>'
            + '<small>Espace réservé au rôle Admin.</small>'
            + '<h2>Comptes</h2><div class="cards">' + cartes + '</div>'
            + '<h2>Par responsabilité</h2><div class="cards">' + roles + '</div>'
            + '<h2>Connexions récentes</h2>'
            + '<table><tr><th>Personne</th><th>Rôle</th><th>Connexion</th>'
              '<th>Durée</th><th>Statut</th></tr>' + lignes_j + '</table>'
            + '<p><a href="/admin/journal">→ Journal complet</a> · '
              '<a href="/admin/export/journal.csv">⬇ Export CSV</a></p>'
            + '<h2>Tentatives de connexion échouées</h2>'
            + '<table><tr><th>Identifiant saisi</th><th>Quand</th><th>Motif</th></tr>'
            + lignes_t + '</table>'
            + '<h2>Transcriptions récentes</h2>'
            + _table_transcriptions(tr)
            + '<h2>Couverture des affectations</h2>'
            + '<table><tr><th>District (avec responsable)</th><th>Responsables</th>'
              '<th>Communes sans superviseur</th></tr>' + lignes_c + '</table>'
            + _pied())


def _options_roles(sel="") -> str:
    return "".join(f'<option value="{ESC(r)}"{" selected" if r==sel else ""}>{ESC(r)}'
                   f'</option>' for r in utilisateurs.RESPONSABILITES)


def _contact_html(u) -> str:
    """Téléphone / CIN / e-mail d'un compte, en lignes (— si tout est vide)."""
    lignes = []
    if u.get("telephone"):
        lignes.append("📞 " + ESC(u["telephone"]))
    if u.get("cin"):
        lignes.append("CIN " + ESC(u["cin"]))
    if u.get("email"):
        lignes.append("✉ " + ESC(u["email"]))
    if u.get("numero_orange_float"):
        lignes.append("Float " + ESC(u["numero_orange_float"]))
    if u.get("sexe"):
        lignes.append("Sexe " + ESC(u["sexe"]))
    return "<br>".join(lignes) if lignes else "—"


def _select_sexe(valeur=None) -> str:
    """<select name="sexe"> avec l'option courante pré-sélectionnée (— vide par défaut)."""
    cur = (valeur or "").strip()
    opts = ['<option value="">—</option>']
    for s in utilisateurs.SEXES:
        sel = ' selected' if s == cur else ''
        opts.append(f'<option value="{s}"{sel}>{s}</option>')
    return '<select name="sexe">' + "".join(opts) + '</select>'


def _options_provinces(provinces) -> str:
    return ('<option value="" selected disabled>— Choisir une province —</option>'
            + "".join(f'<option value="{p["c"]}">{ESC(p["n"])}</option>'
                      for p in provinces))


def page_admin_utilisateurs(conn, message=None, erreur=None) -> str:
    comptes = utilisateurs.lister(conn)
    rows = ""
    for u in comptes:
        aff = utilisateurs.affectation_texte(conn, u["district_affectation"],
                                             u["communes_affectation"],
                                             u.get("districts_affectation"))
        etat = ('<span class="pill ok">actif</span>' if u["actif"]
                else '<span class="pill ko">désactivé</span>')
        bascule = "off" if u["actif"] else "on"
        lg = ESC(u["login"])
        rows += (
            f'<tr><td><b>{lg}</b></td><td>{ESC(u["nom_prenom"])}</td>'
            f'<td>{ESC(u["responsabilite"])}</td><td><small>{ESC(aff)}</small></td>'
            f'<td><small>{_contact_html(u)}</small></td>'
            f'<td>{etat}</td><td>'
            f'<a href="/admin/utilisateurs/modifier?login={urllib.parse.quote(u["login"])}">'
            f'<button type="button" class="sec">Modifier</button></a> '
            f'<form class="inline" method="post" action="/admin/utilisateurs">'
            f'<input type="hidden" name="action" value="actif">'
            f'<input type="hidden" name="login" value="{lg}">'
            f'<input type="hidden" name="etat" value="{bascule}">'
            f'<button class="sec">{"Désactiver" if u["actif"] else "Activer"}</button></form> '
            f'<form class="inline" method="post" action="/admin/utilisateurs" '
            f'onsubmit="return confirm(\'Supprimer {lg} ?\')">'
            f'<input type="hidden" name="action" value="suppr">'
            f'<input type="hidden" name="login" value="{lg}">'
            f'<button class="danger">Supprimer</button></form>'
            f'<form class="inline" method="post" action="/admin/utilisateurs">'
            f'<input type="hidden" name="action" value="reset">'
            f'<input type="hidden" name="login" value="{lg}">'
            f'<input type="password" name="mdp" placeholder="nouveau mot de passe" '
            f'style="width:150px"><button class="sec">Réinit.</button></form>'
            f'</td></tr>')

    msg = f'<div class="msg">{ESC(message)}</div>' if message else ''
    err = f'<div class="err">{erreur}</div>' if erreur else ''   # erreur : HTML permis

    return (_entete("users")
            + '<h1>Gestion des utilisateurs</h1>' + msg + err
            + '<p><a href="/admin/utilisateurs/ajouter"><button>+ Ajouter un '
              'utilisateur</button></a> &nbsp; '
              '<a href="/admin/export/utilisateurs.csv">⬇ Export CSV des comptes</a></p>'
            + '<table><tr><th>Login</th><th>Nom</th><th>Rôle</th><th>Affectation</th>'
              '<th>Contact</th><th>État</th><th>Actions</th></tr>' + rows + '</table>'
            + _pied())


# --- Widgets d'affectation partagés (formulaires Ajouter ET Modifier) -----------
def _geo_roles(conn):
    """(geo_json, groupes_json, prov_opts) pour la cascade Province→Région→District.
    province/région ne sont PAS enregistrées : elles raccourcissent juste les listes."""
    geo = zones.arbre_geo(conn)
    geo["communes"] = zones.communes_par_district(conn)
    geo_json = json.dumps(geo, ensure_ascii=False)
    groupes_json = json.dumps({
        "zone": list(utilisateurs._ROLES_ZONE_ENTIERE),
        "multi": list(utilisateurs._ROLES_MULTI_DISTRICT),
        "un": list(utilisateurs._ROLES_UN_DISTRICT),
        "comm": list(utilisateurs._ROLES_DISTRICT_COMMUNES),
    }, ensure_ascii=False)
    prov_opts = "".join(f'<option value="{p["c"]}">{ESC(p["n"])}</option>'
                        for p in geo["provinces"])
    return geo_json, groupes_json, prov_opts


def _cascade_row(prov_opts, dist_name=None) -> str:
    """Une ligne de cascade Province → Région → District. Seul le select District
    NOMMÉ (dist_name) est soumis ; province/région ne servent qu'à filtrer."""
    dist_attr = f' name="{dist_name}"' if dist_name else ''
    return (
        '<div class="cascade-row" style="display:grid;'
        'grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:6px">'
        f'<select class="c-prov"><option value="">— province —</option>{prov_opts}</select>'
        '<select class="c-reg" disabled><option value="">— région —</option></select>'
        f'<select class="c-dist"{dist_attr} disabled>'
        '<option value="">— district —</option></select></div>')


def _bloc_affectation(prov_opts) -> str:
    """Les 3 widgets d'affectation (montrés/masqués selon le rôle par le JS) :
    1 district / 1 à 5 districts / 1 district + 1 à 7 communes."""
    lignes_dist = "".join(_cascade_row(prov_opts, "district_multi") for _ in range(5))
    comm_selects = "".join(
        f'<select class="comm-cible" name="commune_multi" disabled>'
        f'<option value="">— commune {i} —</option></select>'
        for i in range(1, utilisateurs.MAX_COMMUNES_SUPERVISEUR + 1))
    bloc_communes = (
        '<div id="f-communes" style="grid-column:1/-1">'
        f'<label>District puis communes — 1 à {utilisateurs.MAX_COMMUNES_SUPERVISEUR} '
        '(remplir au moins la 1re commune)</label>'
        '<div class="cc-row" style="display:grid;grid-template-columns:1fr 1fr 1fr;'
        'gap:6px;margin-bottom:6px">'
        f'<select class="cc-prov"><option value="">— province —</option>{prov_opts}</select>'
        '<select class="cc-reg" disabled><option value="">— région —</option></select>'
        '<select class="cc-dist" disabled><option value="">— district —</option></select>'
        '</div>'
        '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));'
        'gap:6px">' + comm_selects + '</div></div>')
    return (
        # 1 district (Traitement / Logistique District / Expert survey)
        '<div id="f-district-single" style="grid-column:1/-1">'
        '<label>District d’affectation</label>' + _cascade_row(prov_opts, "district")
        + '</div>'
        # 1 à 5 districts (Coordonnateur régionale / Comités Techniques)
        + '<div id="f-districts" style="grid-column:1/-1">'
        '<label>Districts d’affectation — 1 à 5 (remplir au moins la 1re ligne)</label>'
        + lignes_dist + '</div>'
        # 1 district + 1 à 7 communes (Superviseur / Logistique Inter-Communale)
        + bloc_communes)


def _cascade_js(geo_json, groupes_json, presel_json="null") -> str:
    """JS du cascade : dépendances Prov→Rég→Dist(→Commune), affichage du bon widget
    selon le rôle, et PRÉ-SÉLECTION (édition) si `presel_json` != null.
    `presel_json` = {"district":d|null,"districts":[..],"communes":[..]}."""
    return (
        '<script>(function(){'
        f'var GEO={geo_json};var ROLES={groupes_json};var PRESEL={presel_json};'
        'function opt(v,t){var o=document.createElement("option");o.value=v;o.textContent=t;return o;}'
        'function fill(sel,list,ph){sel.innerHTML="";sel.appendChild(opt("",ph));'
        '(list||[]).forEach(function(x){sel.appendChild(opt(x.c,x.n));});'
        'sel.disabled=!(list&&list.length);}'
        'function initRow(row){var P=row.querySelector(".c-prov"),'
        'R=row.querySelector(".c-reg"),D=row.querySelector(".c-dist");'
        'P.addEventListener("change",function(){fill(R,GEO.regions[P.value],"— région —");'
        'fill(D,[],"— district —");});'
        'R.addEventListener("change",function(){fill(D,GEO.districts[R.value],"— district —");});'
        '}'
        'document.querySelectorAll(".cascade-row").forEach(initRow);'
        # Section « district + communes » : 1 cascade prov/rég/dist -> remplit les listes de communes.
        'var CP=document.querySelector(".cc-prov"),CR=document.querySelector(".cc-reg"),'
        'CD=document.querySelector(".cc-dist"),COMMS=document.querySelectorAll(".comm-cible");'
        'function resetComms(){COMMS.forEach(function(s,i){fill(s,[],"— commune "+(i+1)+" —");});}'
        'if(CP){'
        'CP.addEventListener("change",function(){fill(CR,GEO.regions[CP.value],"— région —");'
        'fill(CD,[],"— district —");resetComms();});'
        'CR.addEventListener("change",function(){fill(CD,GEO.districts[CR.value],"— district —");resetComms();});'
        'CD.addEventListener("change",function(){COMMS.forEach(function(s,i){'
        'fill(s,GEO.communes[CD.value],"— commune "+(i+1)+" —");});});}'
        'var RS=document.getElementById("ad-role");'
        'function show(id,v){var e=document.getElementById(id);if(e)e.style.display=v?"":"none";}'
        'function majRole(){var r=RS.value;'
        'var multi=ROLES.multi.indexOf(r)>=0,comm=ROLES.comm.indexOf(r)>=0,'
        'un=ROLES.un.indexOf(r)>=0;'
        'show("f-district-single",un);show("f-districts",multi);show("f-communes",comm);}'
        'RS.addEventListener("change",majRole);majRole();'
        # --- Pré-sélection (page de modification) : place les valeurs existantes. ---
        'if(PRESEL){'
        'var d2r={},r2p={};'
        'Object.keys(GEO.regions).forEach(function(p){GEO.regions[p].forEach(function(x){r2p[x.c]=p;});});'
        'Object.keys(GEO.districts).forEach(function(r){GEO.districts[r].forEach(function(x){d2r[x.c]=r;});});'
        'function setDist(row,dc){var reg=d2r[dc];if(reg==null)return;var prov=r2p[reg];'
        'var P=row.querySelector(".c-prov"),R=row.querySelector(".c-reg"),D=row.querySelector(".c-dist");'
        'P.value=prov;fill(R,GEO.regions[prov],"— région —");R.value=reg;'
        'fill(D,GEO.districts[reg],"— district —");D.value=dc;}'
        'var rr=RS.value;'
        'if(ROLES.un.indexOf(rr)>=0&&PRESEL.district!=null){'
        'setDist(document.querySelector("#f-district-single .cascade-row"),PRESEL.district);}'
        'if(ROLES.multi.indexOf(rr)>=0){var rows=document.querySelectorAll("#f-districts .cascade-row");'
        '(PRESEL.districts||[]).forEach(function(dc,i){if(rows[i])setDist(rows[i],dc);});}'
        'if(ROLES.comm.indexOf(rr)>=0&&PRESEL.district!=null){var reg=d2r[PRESEL.district],prov=r2p[reg];'
        'CP.value=prov;fill(CR,GEO.regions[prov],"— région —");CR.value=reg;'
        'fill(CD,GEO.districts[reg],"— district —");CD.value=PRESEL.district;'
        'COMMS.forEach(function(s,i){fill(s,GEO.communes[PRESEL.district],"— commune "+(i+1)+" —");});'
        '(PRESEL.communes||[]).forEach(function(cc,i){if(COMMS[i])COMMS[i].value=cc;});}'
        '}'
        '})();</script>')


_AIDE_AFFECTATION = (
    '<div class="note"><b>Affectation selon le rôle</b> — '
    'Admin / Coordonnateur Nationale : toute la zone (rien à choisir) · '
    'Coordonnateur régionale, Comités Techniques : 1 à 5 districts · '
    'Traitement, Logistique District, Expert survey : 1 district · '
    f'Superviseur Technique, Logistique Inter-Communale : 1 district + 1 à '
    f'{utilisateurs.MAX_COMMUNES_SUPERVISEUR} communes.</div>')


def page_admin_ajouter(conn, message=None, erreur=None) -> str:
    """Formulaire d'ajout d'un utilisateur (cascade Province→Région→District→Commune)
    + import Excel. Page séparée de la liste (page_admin_utilisateurs)."""
    msg = f'<div class="msg">{ESC(message)}</div>' if message else ''
    err = f'<div class="err">{erreur}</div>' if erreur else ''   # erreur : HTML permis
    geo_json, groupes_json, prov_opts = _geo_roles(conn)

    form_ajout = (
        '<form method="post" action="/admin/utilisateurs" class="grid-form">'
        '<input type="hidden" name="action" value="ajouter">'
        '<div><label>Nom et prénom</label><input name="nom_prenom" required></div>'
        '<div><label>Responsabilité</label>'
        f'<select id="ad-role" name="responsabilite">{_options_roles()}</select></div>'
        '<div><label>Téléphone</label>'
        '<input name="telephone" inputmode="tel" placeholder="0340000000"></div>'
        '<div><label>CIN (12 chiffres)</label>'
        '<input name="cin" inputmode="numeric" placeholder="101012345678"></div>'
        '<div><label>Adresse e-mail</label>'
        '<input type="email" name="email" placeholder="nom@example.mg"></div>'
        '<div><label>N° Orange (Float)</label>'
        '<input name="numero_orange_float" inputmode="tel" '
        'placeholder="0320000000"></div>'
        '<div><label>Sexe</label>' + _select_sexe() + '</div>'
        '<div><label>Login</label><input name="login" required></div>'
        '<div><label>Mot de passe</label><input type="password" name="mot_de_passe" required></div>'
        + _bloc_affectation(prov_opts)
        + '<div style="align-self:end;grid-column:1/-1">'
        '<button>Ajouter l’utilisateur</button></div>'
        '</form>')

    form_import = (
        '<form method="post" action="/admin/import" enctype="multipart/form-data" '
        'class="grid-form">'
        '<div><label>Fichier Excel (.xlsx)</label>'
        '<input type="file" name="fichier" accept=".xlsx" required></div>'
        '<div style="align-self:end"><button>Importer</button></div></form>'
        '<div class="note">Colonnes : <b>nom_prenom, responsabilite, telephone, cin, '
        'email, numero_orange_float, sexe, login, mot_de_passe, district, communes, '
        'districts</b> '
        '(telephone, cin, email, numero_orange_float, sexe facultatifs). '
        '<a href="/admin/modele.xlsx">⬇ Télécharger le modèle</a>. '
        '⚠️ Le fichier contient des mots de passe en clair : <b>supprimez-le</b> '
        'après l’import.</div>')

    return (_entete("users_add")
            + '<h1>Ajouter un utilisateur</h1>'
            + '<p><a href="/admin/utilisateurs">← Retour à la liste</a></p>'
            + msg + err + _AIDE_AFFECTATION
            + form_ajout
            + '<h2>Importer depuis Excel</h2>' + form_import
            + _cascade_js(geo_json, groupes_json)
            + _pied())


def page_admin_modifier(conn, login, message=None, erreur=None) -> str:
    """Formulaire de MODIFICATION d'un utilisateur, pré-rempli avec ses infos
    existantes. Le login (clé) est affiché mais non modifiable ; le mot de passe est
    facultatif (vide = inchangé). L'affectation courante est pré-sélectionnée (JS)."""
    u = utilisateurs.obtenir(conn, login)
    if u is None:
        return (_entete("users") + '<h1>Modifier un utilisateur</h1>'
                + f'<div class="err">Login inconnu : {ESC(login)}.</div>'
                + '<p><a href="/admin/utilisateurs">← Retour à la liste</a></p>'
                + _pied())
    msg = f'<div class="msg">{ESC(message)}</div>' if message else ''
    err = f'<div class="err">{erreur}</div>' if erreur else ''
    geo_json, groupes_json, prov_opts = _geo_roles(conn)
    presel = json.dumps({
        "district": u["district_affectation"],
        "districts": u.get("districts_affectation") or [],
        "communes": u.get("communes_affectation") or [],
    }, ensure_ascii=False)
    lg = ESC(u["login"])

    def champ(nom, valeur, **attrs):
        a = "".join(f' {k}="{v}"' for k, v in attrs.items())
        return f'<input name="{nom}" value="{ESC(valeur or "")}"{a}>'

    form = (
        '<form method="post" action="/admin/utilisateurs" class="grid-form">'
        '<input type="hidden" name="action" value="modifier">'
        f'<input type="hidden" name="login" value="{lg}">'
        '<div><label>Nom et prénom</label>'
        + champ("nom_prenom", u["nom_prenom"], required="required") + '</div>'
        '<div><label>Responsabilité</label>'
        f'<select id="ad-role" name="responsabilite">'
        f'{_options_roles(u["responsabilite"])}</select></div>'
        '<div><label>Téléphone</label>'
        + champ("telephone", u.get("telephone"), inputmode="tel") + '</div>'
        '<div><label>CIN (12 chiffres)</label>'
        + champ("cin", u.get("cin"), inputmode="numeric") + '</div>'
        '<div><label>Adresse e-mail</label>'
        + champ("email", u.get("email"), type="email") + '</div>'
        '<div><label>N° Orange (Float)</label>'
        + champ("numero_orange_float", u.get("numero_orange_float"),
                inputmode="tel") + '</div>'
        '<div><label>Sexe</label>' + _select_sexe(u.get("sexe")) + '</div>'
        f'<div><label>Login (non modifiable)</label>'
        f'<input value="{lg}" readonly style="background:#eee"></div>'
        '<div><label>Nouveau mot de passe <small>(vide = inchangé)</small></label>'
        '<input type="password" name="mot_de_passe" autocomplete="new-password"></div>'
        + _bloc_affectation(prov_opts)
        + '<div style="align-self:end;grid-column:1/-1">'
        '<button>Enregistrer les modifications</button></div>'
        '</form>')

    return (_entete("users")
            + f'<h1>Modifier « {lg} »</h1>'
            + '<p><a href="/admin/utilisateurs">← Retour à la liste</a></p>'
            + msg + err + _AIDE_AFFECTATION
            + form
            + _cascade_js(geo_json, groupes_json, presel)
            + _pied())


def page_journal(conn) -> str:
    jr = journal.journal(conn, limite=500)
    lignes = "".join(
        f'<tr><td>{ESC(x["login"] or "—")}</td><td>{ESC(x["nom_prenom"] or "—")}</td>'
        f'<td>{ESC(x["responsabilite"] or "—")}</td>'
        f'<td>{ESC(x["connexion_le"] or "—")}</td>'
        f'<td>{ESC(x["deconnexion_le"] or "—")}</td><td>{ESC(x["duree"])}</td>'
        f'<td><span class="pill {"ok" if x["statut"]=="terminée" else "warn"}">'
        f'{x["statut"]}</span></td><td>{ESC(x["ip"] or "")}</td></tr>' for x in jr) \
        or '<tr><td colspan="8"><small>Aucune connexion.</small></td></tr>'
    return (_entete("journal")
            + '<h1>Journal de connexion</h1>'
            + '<p><a href="/admin/export/journal.csv">⬇ Export CSV</a></p>'
            + '<table><tr><th>Login</th><th>Nom</th><th>Rôle</th><th>Connexion</th>'
              '<th>Déconnexion</th><th>Durée</th><th>Statut</th><th>IP</th></tr>'
            + lignes + '</table>' + _pied())
